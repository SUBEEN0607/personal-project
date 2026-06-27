"""
데이터 입력 표준화 모듈
- 컬럼명 자동 매핑 (한글/영문/약어 대응)
- 단위 자동 감지 및 변환 (원/백만원/억원)
- 날짜 형식 자동 파싱
- 유효성 검증 + 오류 안내
"""
import pandas as pd
import re

# 컬럼명 자동 매핑 테이블
_COL_MAP = {
    "회사명": ["회사명", "company", "기업명", "투자처", "피투자사", "포트폴리오사", "name", "corp_name", "종목명",
              "회사", "법인명", "기업", "investee", "portfolio_company", "comp_name", "company_name"],
    "섹터": ["섹터", "sector", "업종", "산업", "분야", "industry", "투자분야",
            "업종분류", "산업분류", "카테고리", "category", "segment", "투자업종"],
    "투자단계": ["투자단계", "stage", "라운드", "round", "투자라운드", "series", "단계",
               "투자차수", "차수", "투자round", "funding_round", "investment_stage", "시리즈"],
    "투자일": ["투자일", "투자일자", "invest_date", "investment_date", "투자실행일", "집행일",
             "투자실행일자", "집행일자", "최초투자일", "첫투자일", "date_invested", "closing_date"],
    "기준일": ["기준일", "평가일", "base_date", "valuation_date", "evaluation_date", "보고기준일",
             "평가기준일", "보고일", "reporting_date", "nav_date", "기준일자"],
    "투자금액_백만원": ["투자금액_백만원", "투자금액", "투자원금", "investment", "invested", "inv_amount",
                      "투자금액(백만)", "투자금액(백만원)", "투자액", "출자금",
                      "투자원금(백만원)", "투자원금(백만)", "cost", "invested_amount", "투자비용",
                      "취득원가", "취득금액", "commitment", "drawn_amount", "납입금액"],
    "현재가치_백만원": ["현재가치_백만원", "현재가치", "평가가치", "fair_value", "current_value", "fv",
                      "현재가치(백만)", "현재가치(백만원)", "평가액", "잔존가치",
                      "공정가치", "nav", "net_asset_value", "시가", "평가금액", "fmv",
                      "fair_market_value", "remaining_value", "미실현가치"],
    "회수금액_백만원": ["회수금액_백만원", "회수금액", "회수액", "distribution", "realized", "회수",
                      "회수금액(백만)", "회수금액(백만원)", "배분금액", "분배금",
                      "실현수익", "회수원금", "배당금", "분배금액", "realized_amount",
                      "proceeds", "exit_amount", "실현금액", "현금회수"],
    "지분율_%": ["지분율_%", "지분율", "stake", "ownership", "지분", "보유지분", "지분율(%)",
               "소유지분", "지분비율", "ownership_pct", "equity_stake", "share_pct", "보유비율"],
}

# 필수 컬럼
_REQUIRED = ["회사명", "섹터", "투자단계", "투자일", "기준일", "투자금액_백만원", "회수금액_백만원"]
_OPTIONAL = ["현재가치_백만원", "지분율_%"]


def _normalize(s: str) -> str:
    return re.sub(r"[^가-힣a-z]", "", s.strip().lower())

def _find_best_match(col_name: str) -> str | None:
    col_clean = col_name.strip().lower().replace(" ", "").replace("_", "")
    col_norm = _normalize(col_name)
    # 1차: 정확 매칭
    for standard, variants in _COL_MAP.items():
        for v in variants:
            if col_clean == v.lower().replace(" ", "").replace("_", ""):
                return standard
    # 2차: 부분 매칭
    for standard, variants in _COL_MAP.items():
        for v in variants:
            v_clean = v.lower().replace(" ", "").replace("_", "")
            if v_clean in col_clean or col_clean in v_clean:
                return standard
    # 3차: 정규화 후 핵심 키워드 매칭
    keyword_map = {
        "회사명": ["회사", "기업", "company", "investee"],
        "섹터": ["섹터", "업종", "산업", "sector"],
        "투자단계": ["단계", "라운드", "stage", "round", "series"],
        "투자일": ["투자일", "집행일", "invest"],
        "기준일": ["기준일", "평가일", "보고일"],
        "투자금액_백만원": ["투자금", "투자원금", "출자", "invest", "cost", "commitment"],
        "현재가치_백만원": ["현재가치", "평가가치", "공정가치", "fair", "nav", "fmv"],
        "회수금액_백만원": ["회수", "분배", "배분", "distribut", "realized", "proceeds"],
        "지분율_%": ["지분", "stake", "ownership"],
    }
    for standard, keywords in keyword_map.items():
        for kw in keywords:
            if kw in col_norm or kw in col_clean:
                return standard
    return None


def _detect_unit(series: pd.Series) -> str:
    """숫자 컬럼의 단위 추정 (원/백만원/억원)"""
    vals = series.dropna()
    if len(vals) == 0:
        return "백만원"
    median = vals.median()
    if median > 1_000_000_000:
        return "원"
    elif median > 100_000:
        return "백만원"
    elif median > 100:
        return "억원"
    else:
        return "백만원"


def _convert_to_백만원(series: pd.Series, unit: str) -> pd.Series:
    if unit == "원":
        return series / 1_000_000
    elif unit == "억원":
        return series * 100
    return series


def standardize(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """
    입력 DataFrame을 표준 형식으로 변환.
    Returns: (표준화된 df, 경고 메시지 리스트)
    """
    warnings = []
    result = pd.DataFrame()

    # 1. 컬럼명 자동 매핑
    mapped = {}
    unmapped = []
    for col in df.columns:
        match = _find_best_match(col)
        if match:
            mapped[col] = match
        else:
            unmapped.append(col)

    if unmapped:
        warnings.append(f"매핑 안 된 컬럼 (무시됨): {', '.join(unmapped)}")

    df_renamed = df.rename(columns=mapped)

    # 2. 필수 컬럼 확인
    missing = [c for c in _REQUIRED if c not in df_renamed.columns]
    if missing:
        warnings.append(f"필수 컬럼 누락: {', '.join(missing)}")
        return df, warnings

    # 3. 선택 컬럼 기본값
    if "현재가치_백만원" not in df_renamed.columns:
        df_renamed["현재가치_백만원"] = 0
        warnings.append("현재가치 컬럼 없음 → 0으로 설정 (자동 밸류에이션 사용 가능)")
    if "지분율_%" not in df_renamed.columns:
        df_renamed["지분율_%"] = 10.0
        warnings.append("지분율 컬럼 없음 → 10%로 기본 설정")

    # 4. 숫자 컬럼 단위 자동 변환
    for col in ["투자금액_백만원", "현재가치_백만원", "회수금액_백만원"]:
        if col in df_renamed.columns:
            df_renamed[col] = pd.to_numeric(df_renamed[col], errors="coerce").fillna(0)
            unit = _detect_unit(df_renamed[col])
            if unit != "백만원":
                df_renamed[col] = _convert_to_백만원(df_renamed[col], unit)
                warnings.append(f"{col}: {unit} 단위 감지 → 백만원으로 자동 변환")

    # 5. 날짜 자동 파싱
    for col in ["투자일", "기준일"]:
        if col in df_renamed.columns:
            try:
                df_renamed[col] = pd.to_datetime(df_renamed[col], infer_datetime_format=True)
            except Exception:
                warnings.append(f"{col} 날짜 형식 인식 실패 — YYYY-MM-DD 형식으로 입력해주세요")

    # 6. 데이터 유효성 검증
    if (df_renamed["투자금액_백만원"] <= 0).any():
        bad = df_renamed[df_renamed["투자금액_백만원"] <= 0]["회사명"].tolist()
        warnings.append(f"투자금액 0 이하: {', '.join(str(x) for x in bad)}")

    if df_renamed["회사명"].duplicated().any():
        dups = df_renamed[df_renamed["회사명"].duplicated()]["회사명"].tolist()
        warnings.append(f"중복 회사명: {', '.join(str(x) for x in dups)}")

    # 표준 컬럼만 추출
    std_cols = _REQUIRED + _OPTIONAL
    available = [c for c in std_cols if c in df_renamed.columns]
    result = df_renamed[available].copy()

    return result, warnings


def generate_guide_excel() -> bytes:
    """표준 입력 가이드 Excel 생성 (3시트)"""
    import io
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        sample = pd.DataFrame({
            "회사명": ["넥스틸바이오", "코리아로지텍", "그린솔라원", "미래모빌리티", "케어에이아이"],
            "섹터": ["바이오", "SaaS/물류", "신재생에너지", "모빌리티", "의료AI"],
            "투자단계": ["Series B", "Series A", "Pre-A", "Series C", "Series B"],
            "투자일": ["2020-04-10", "2021-02-20", "2020-08-15", "2021-11-01", "2020-06-30"],
            "기준일": ["2024-06-30"] * 5,
            "투자금액_백만원": [1500, 600, 300, 4000, 1200],
            "현재가치_백만원": [5100, 1560, 0, 3200, 3840],
            "회수금액_백만원": [0, 0, 960, 0, 0],
            "지분율_%": [8.5, 12.0, 15.0, 5.0, 10.0],
        })
        sample.to_excel(w, sheet_name="샘플데이터", index=False)

        guide = pd.DataFrame({
            "컬럼명": list(_COL_MAP.keys()),
            "필수": ["필수"]*7 + ["선택"]*2,
            "설명": [
                "포트폴리오 회사명 (법인명 권장)", "투자 섹터", "투자 라운드 단계",
                "최초 투자 실행일 (YYYY-MM-DD)", "평가 기준일 (YYYY-MM-DD)",
                "투자 원금 (백만원)", "현재 평가가치 (0이면 자동 추정)",
                "회수한 금액 (미회수 시 0)", "취득 지분율 (미입력 시 10%)",
            ],
            "자동 인식 가능한 다른 이름": [
                ", ".join(v[1:]) for v in _COL_MAP.values()
            ],
        })
        guide.to_excel(w, sheet_name="입력가이드", index=False)

        sectors = pd.DataFrame({
            "섹터명": ["바이오","의료AI","SaaS","SaaS/물류","모빌리티","자율주행","에듀테크",
                      "신재생에너지","딥테크","AI","반도체","핀테크","커머스","콘텐츠","게임",
                      "애그테크","푸드테크","물류","ESG","헬스케어"],
            "P/S 배수": [8.0,7.0,5.0,4.5,3.0,5.0,3.5,2.5,5.0,7.0,6.0,4.5,2.0,3.0,4.0,3.0,2.5,2.0,2.0,4.0],
        })
        sectors.to_excel(w, sheet_name="섹터목록", index=False)

        wb = w.book
        gf = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
        wf = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
        for sn in wb.sheetnames:
            ws = wb[sn]
            for cell in ws[1]:
                cell.fill = gf; cell.font = wf
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                cl = col[0].column_letter
                mx = max(sum(2 if ord(c) > 127 else 1 for c in str(cell.value or "")) for cell in col)
                ws.column_dimensions[cl].width = min(mx + 4, 50)

    return buf.getvalue()
