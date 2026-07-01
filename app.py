import streamlit as st
import pandas as pd
import io
import plotly.express as px
import plotly.graph_objects as go
import base64, os, re, requests

from calculator import load_portfolio, run_all, portfolio_summary
from rag import answer_question
from irr import j_curve_data
from simulator import simulate_exit, optimal_exit_timing
from db import save_snapshot, load_quarters, load_trend
from report import generate_full_pdf
from dart_client import search_company, get_financials
from benchmark import (
    get_base_rate, get_exchange_rate, get_vc_trend,
    get_kvic_sector_summary, get_kvic_yearly_trend,
)
from commentary import (
    generate_commentary,
    interpret_jcurve,
    interpret_scenario,
    interpret_quarterly_trend,
    interpret_dart_financials,
    interpret_macro,
)



st.set_page_config(page_title="PE/VC 분기 보고 도우미", layout="wide")

# ── 표지/앱 전환 상태 ─────────────────────────────
if "show_cover" not in st.session_state:
    st.session_state["show_cover"] = True

st.markdown("""
<style>
@import url('https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/static/pretendard.css');

/* ── Design System ── */
html, body, [class*="css"], .stApp {
    font-family: 'Pretendard', -apple-system, system-ui, sans-serif !important;
    background-color: #ffffff !important;
    color: #1a1a1a !important;
}
[data-testid="stAppViewContainer"] { background-color: #ffffff !important; }

/* Sidebar */
[data-testid="stSidebar"] {
    background-color: #f8f8f6 !important;
    border-right: 1px solid #eee !important;
    border-left: none !important;
}
[data-testid="stSidebar"] h1, [data-testid="stSidebar"] h2, [data-testid="stSidebar"] h3 {
    font-size: 14px !important;
}

/* Typography — tight, bold headings */
h1 { font-size: 32px !important; font-weight: 800 !important; letter-spacing: -0.04em !important; }
h2 { font-size: 22px !important; font-weight: 700 !important; letter-spacing: -0.03em !important; }
h3 { font-size: 18px !important; font-weight: 700 !important; letter-spacing: -0.02em !important; }
h4 { font-size: 15px !important; font-weight: 600 !important; letter-spacing: -0.01em !important; }
h1,h2,h3,h4,h5,h6 { color: #1a1a1a !important; }

/* Cards */
[data-testid="stVerticalBlockBorderWrapper"] {
    border: 1px solid #e5e5e5 !important;
    border-radius: 12px !important;
    box-shadow: none !important;
    background: #ffffff !important;
}
[data-testid="stVerticalBlockBorderWrapper"]:hover {
    transform: none !important;
    box-shadow: 0 2px 8px rgba(0,0,0,0.04) !important;
}

/* Metrics */
[data-testid="stMetric"] {
    background: #ffffff;
    border: 1px solid #e5e5e5;
    border-radius: 12px;
    padding: 20px 24px;
    box-shadow: none;
}
[data-testid="stMetricLabel"] {
    font-size: 10px !important; color: #999 !important;
    letter-spacing: 0.1em !important; text-transform: uppercase !important;
    font-weight: 500 !important;
}
[data-testid="stMetricValue"] {
    font-size: 30px !important; font-weight: 800 !important;
    color: #1a1a1a !important; letter-spacing: -0.04em !important;
}
[data-testid="stMetricDelta"] { font-size: 12px !important; font-weight: 500 !important; }

/* Tabs — minimal underline */
[data-baseweb="tab-list"] { border-bottom: 1px solid #eae8e4 !important; gap: 0 !important; }
[data-baseweb="tab"] {
    font-size: 13px !important; font-weight: 500 !important;
    color: #999 !important; padding: 12px 20px !important;
    border-radius: 0 !important; letter-spacing: 0.01em !important;
}
[aria-selected="true"][data-baseweb="tab"] {
    color: #1a1a1a !important;
    border-bottom: 2px solid #1b5e20 !important;
    background-color: transparent !important;
    font-weight: 600 !important;
}

/* Buttons — clean, subtle */
.stButton > button {
    color: #1a1a1a !important;
    border: 1px solid #ddd !important;
    background-color: #ffffff !important;
    border-radius: 8px !important;
    padding: 8px 20px !important;
    font-size: 13px !important;
    font-weight: 500 !important;
    transition: all 0.15s ease !important;
    box-shadow: none !important;
}
.stButton > button:hover {
    border-color: #1b5e20 !important;
    color: #1b5e20 !important;
    background-color: #f5f9f5 !important;
}
.stButton > button[kind="primary"],
.stButton > button[data-testid="stBaseButton-primary"] {
    background-color: #ffffff !important; border: 1.5px solid #1b5e20 !important;
    color: #1a1a1a !important;
    border: none !important;
    border-radius: 24px !important;
    padding: 10px 36px !important;
    font-weight: 600 !important;
    letter-spacing: 0.03em !important;
}
.stButton > button[kind="primary"]:hover,
.stButton > button[data-testid="stBaseButton-primary"]:hover {
    background-color: #d4c4a8 !important;
    color: #1a1a1a !important;
}

/* Download button — soft green */
.stDownloadButton > button {
    background-color: rgba(27,94,32,0.08) !important;
    color: #1b5e20 !important;
    border: 1px solid rgba(27,94,32,0.2) !important;
    border-radius: 8px !important;
    font-weight: 500 !important;
}
.stDownloadButton > button:hover {
    background-color: rgba(27,94,32,0.15) !important;
    color: #1b5e20 !important;
}

/* DataFrames — green header */
[data-testid="stDataFrame"] thead th, [data-testid="stDataFrame"] th {
    background-color: #1b5e20 !important;
    color: #ffffff !important;
    font-weight: 600 !important;
    font-size: 11px !important;
    letter-spacing: 0.04em !important;
    text-transform: uppercase !important;
    text-align: center !important;
    padding: 10px 14px !important;
    border: none !important;
}
[data-testid="stDataFrame"] tbody td {
    text-align: center !important;
    padding: 10px 14px !important;
    border-bottom: 1px solid #eee !important;
    font-size: 13px !important;
    background-color: #ffffff !important;
}
[data-testid="stDataFrame"] tbody tr:hover td { background-color: #f0f7f0 !important; }

/* Inputs — subtle */
.stTextInput > div > div > input,
.stNumberInput > div > div > input {
    border: 1px solid #ddd !important;
    border-radius: 8px !important;
    padding: 10px 14px !important;
    font-size: 14px !important;
    background-color: #ffffff !important;
}
.stTextInput > div > div > input:focus,
.stNumberInput > div > div > input:focus {
    border-color: #1b5e20 !important;
    box-shadow: 0 0 0 2px rgba(27,94,32,0.08) !important;
}
[data-baseweb="input"]:focus-within {
    border-color: #1b5e20 !important;
    box-shadow: 0 0 0 2px rgba(27,94,32,0.08) !important;
}

/* Selectbox, Slider */
[data-testid="stSelectbox"] > div > div { border: 1px solid #ddd !important; border-radius: 8px !important; }
[data-testid="stSlider"] [role="slider"] { color: #1a1a1a !important; }
[data-baseweb="slider"] [role="slider"] { background-color: #1a1a1a !important; border-color: #1a1a1a !important; }
[data-baseweb="slider"] [role="progressbar"] > div:first-child { background-color: #1a1a1a !important; }
[data-baseweb="slider"] [role="progressbar"] { background-color: #e0e0e0 !important; }

/* Expander */
[data-testid="stExpander"] { border: 1px solid #e5e5e5 !important; border-radius: 12px !important; background: #ffffff !important; }

/* Divider */
hr { border: none !important; border-top: 1px solid #ddd !important; margin: 20px 0 !important; }

/* Alert */
.stAlert > div { border-radius: 10px !important; }

/* Hide clutter */
#MainMenu { visibility: hidden; }
footer { visibility: hidden; }

/* Cover button */
.cover-btn > button {
    background-color: transparent !important;
    color: #ffffff !important;
    border: 1.5px solid rgba(255,255,255,0.5) !important;
    border-radius: 24px !important;
    padding: 12px 36px !important;
    font-size: 14px !important;
    font-weight: 500 !important;
    letter-spacing: 0.05em !important;
}
.cover-btn > button:hover {
    background-color: rgba(255,255,255,0.12) !important;
    border-color: #ffffff !important;
    color: #ffffff !important;
}

/* Custom card class */
.clean-card {
    background: #ffffff;
    border: 1px solid #eae8e4;
    border-radius: 10px;
    padding: 24px;
}
.stat-large {
    font-size: 48px;
    font-weight: 800;
    letter-spacing: -0.04em;
    line-height: 1;
    color: #1a1a1a;
}
.stat-label {
    font-size: 10px;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    color: #999;
    font-weight: 500;
    margin-bottom: 6px;
}
.stat-desc {
    font-size: 12px;
    color: #999;
    margin-top: 8px;
}
.accent { color: #1b5e20; }
.section-label {
    font-size: 10px;
    letter-spacing: 0.15em;
    text-transform: uppercase;
    color: #bbb;
    font-weight: 600;
    margin-bottom: 4px;
}
</style>
""", unsafe_allow_html=True)

# ── 배경 이미지 base64 로드 ───────────────────────
_wp_b64 = ""
for _try_path in [
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "skku_wallpaper.jpg"),
    os.path.join(os.getcwd(), "skku_wallpaper.jpg"),
    r"C:\Users\Lenovo\personal-project\skku_wallpaper.jpg",
]:
    if os.path.exists(_try_path):
        with open(_try_path, "rb") as _f:
            _wp_b64 = base64.b64encode(_f.read()).decode()
        break

if st.session_state["show_cover"]:
    _b64_src = f"data:image/jpeg;base64,{_wp_b64}" if _wp_b64 else ""

    st.markdown(f"""
    <style>
    [data-testid="stAppViewContainer"] {{ background: transparent !important; }}
    [data-testid="stMain"] {{ background: transparent !important; }}
    .block-container {{ padding: 0 !important; max-width: 100% !important; }}
    [data-testid="stMainBlockContainer"] {{ padding: 0 !important; max-width: 100% !important; }}
    section[data-testid="stMain"] > div {{ padding: 0 !important; }}
    .stMainBlockContainer {{ max-width: 100% !important; padding: 0 !important; }}
    html, body, .stApp {{ background: transparent !important; }}
    [data-testid="stSidebar"] {{ display: none !important; }}
    [data-testid="stHeader"] {{ display: none !important; }}
    section[data-testid="stMain"] > div {{ padding: 0 !important; }}
    .block-container {{ padding: 0 !important; max-width: 100% !important; }}

    .cv-full {{
        position: fixed; top: 0; left: 0; width: 100vw; height: 100vh;
        overflow: hidden; z-index: 0; pointer-events: none;
    }}
    .cv-bg {{
        position: absolute; top: 0; left: 0; width: 100%; height: 100%;
        background: {"url('" + _b64_src + "') center/cover no-repeat" if _b64_src else "linear-gradient(135deg,#2d7a35,#1b5e20)"};
    }}
    .cv-overlay {{
        position: absolute; top: 0; left: 0; width: 100%; height: 100%;
        background: linear-gradient(180deg,
            rgba(15,50,20,0.3) 0%,
            rgba(20,60,25,0.45) 40%,
            rgba(15,45,18,0.6) 65%,
            rgba(10,30,12,0.8) 100%);
    }}
    .cv-content {{
        position: relative; width: 100%; min-height: 100vh;
        display: flex; flex-direction: column; align-items: center; justify-content: center;
        text-align: center; z-index: 10; padding: 40px 20px;
    }}
    .cv-top {{ font-size:12px; color:rgba(255,255,255,0.7); letter-spacing:0.3em;
        text-transform:uppercase; margin-bottom:40px; font-weight:600;
        text-shadow:0 2px 10px rgba(0,0,0,0.5); }}
    .cv-title {{ font-size:clamp(80px,12vw,150px); font-weight:900; line-height:0.88;
        letter-spacing:-0.04em; color:#fff;
        text-shadow:0 0 80px rgba(255,255,255,0.15), 0 6px 30px rgba(0,0,0,0.5); }}
    .cv-sub {{ font-size:clamp(24px,3.5vw,42px); font-weight:700; letter-spacing:0.08em;
        margin-top:10px; color:rgba(255,255,255,0.9);
        text-shadow:0 4px 15px rgba(0,0,0,0.5); }}
    .cv-tags {{ display:flex; gap:10px; flex-wrap:wrap; justify-content:center;
        margin-top:50px; margin-bottom:25px; }}
    .cv-tag {{ font-size:10px; color:rgba(255,255,255,0.55); letter-spacing:0.06em;
        padding:4px 14px; border:1px solid rgba(255,255,255,0.2); border-radius:20px; }}
    .cv-name {{ font-size:12px; color:rgba(255,255,255,0.55); letter-spacing:0.08em;
        font-weight:500; text-shadow:0 2px 8px rgba(0,0,0,0.4); }}
    .cv-enter {{ margin-top:30px; }}
    .cv-enter button {{
        background: transparent !important; color: #fff !important;
        border: 1.5px solid rgba(255,255,255,0.4) !important;
        border-radius: 24px !important; padding: 10px 40px !important;
        font-size: 13px !important; letter-spacing: 0.05em !important;
        cursor: pointer !important;
    }}
    .cv-enter button:hover {{
        background: rgba(255,255,255,0.1) !important;
        border-color: rgba(255,255,255,0.6) !important;
    }}
    </style>

    <div class="cv-full">
        <div class="cv-bg"></div>
        <div class="cv-overlay"></div>
        <div class="cv-content">
            <div class="cv-top">SDIC &middot; SKKU Digital IT Consulting</div>
            <div class="cv-title">PE/ VC</div>
            <div class="cv-sub">분기 보고 도우미</div>
            <div class="cv-name" style="margin-top:16px;">이수빈 &middot; 개인 프로젝트</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown(
        '<div style="margin-top:500px; position:relative; z-index:100; padding-bottom:10px;">',
        unsafe_allow_html=True,
    )
    col_l, col_c, col_r = st.columns([2, 1, 2])
    with col_c:
        if st.button("Enter →", use_container_width=True, type="primary"):
            st.session_state["show_cover"] = False
            st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('<div style="text-align:center;padding-bottom:40px;position:relative;z-index:100;"><span style="font-size:12px;color:rgba(255,255,255,0.45);">데이터 수집부터 성과 분석, LP 보고서와 IC 장표 작성까지 — 분기 보고에 필요한 모든 과정을 하나의 화면에서 완성합니다.</span></div>', unsafe_allow_html=True)

    st.stop()

# ── 일반 헤더: 표지 통과 후 ──────────────────────
st.markdown("""
<div style="padding: 0 0 20px 0; border-bottom: 1px solid #e5e5e5; margin-bottom: 24px;">
  <div style="display:flex; align-items:baseline; gap:12px;">
    <span style="font-size:24px; font-weight:800; color:#1a1a1a; letter-spacing:-0.04em;">PE/VC</span>
    <span style="font-size:13px; font-weight:400; color:#bbb;">분기 보고 도우미</span>
  </div>
</div>
""", unsafe_allow_html=True)

# ── 사이드바 ─────────────────────────────────────
with st.sidebar:
    st.markdown("""
<div style="padding: 12px 0 16px 0; border-bottom: 1px solid #eae8e4; margin-bottom: 20px;">
  <div style="font-size: 20px; font-weight: 800; color: #1a1a1a; letter-spacing:-0.04em;">PE/VC</div>
  <div style="font-size: 10px; color: #bbb; letter-spacing: 0.1em; font-weight: 500; margin-top: 2px;">SDIC · 이수빈</div>
</div>
""", unsafe_allow_html=True)

    st.markdown('<p style="font-size:12px;letter-spacing:0.1em;text-transform:uppercase;color:#1b5e20;font-weight:700;margin-bottom:6px;">펀드 정보</p>', unsafe_allow_html=True)
    fund_name = st.text_input("펀드명", value="SDIC 성장투자 1호")
    fund_strategy = st.selectbox("종류", ["VC (벤처캐피탈)", "PE (사모펀드)"])

    st.markdown("---")
    st.markdown('<p style="font-size:12px;letter-spacing:0.1em;text-transform:uppercase;color:#1b5e20;font-weight:700;margin-bottom:6px;">데이터</p>', unsafe_allow_html=True)
    st.markdown('<p style="font-size:10px;color:#999;line-height:1.5;margin-bottom:8px;">엑셀 또는 CSV 파일을 업로드하세요. 다양한 양식을 자동으로 인식합니다.</p>', unsafe_allow_html=True)
    uploaded = st.file_uploader("CSV / Excel", type=["csv", "xlsx"], label_visibility="collapsed")
    use_sample = st.button("샘플 불러오기", use_container_width=True)
    if st.button("입력 가이드", use_container_width=True):
        from data_parser import generate_guide_excel
        guide_buf = generate_guide_excel()
        st.download_button("다운로드", guide_buf, file_name="PE_VC_입력_가이드.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)

    st.markdown("---")
    st.markdown('<p style="font-size:12px;letter-spacing:0.1em;text-transform:uppercase;color:#1b5e20;font-weight:700;margin-bottom:6px;">분기 저장</p>', unsafe_allow_html=True)
    quarter = st.text_input("분기 (예: 2024Q1)", value="2024Q1")
    if st.button("현재 데이터 저장"):
        if "result_df" in st.session_state:
            save_snapshot(st.session_state["result_df"], quarter)
            st.success(f"{quarter} 저장 완료")
        else:
            st.warning("먼저 데이터를 로드하세요.")

if uploaded:
    from data_parser import standardize
    if uploaded.name.endswith(".xlsx"):
        xl = pd.ExcelFile(uploaded)
        sheet = st.sidebar.selectbox("시트 선택", xl.sheet_names)
        raw_input = pd.read_excel(uploaded, sheet_name=sheet)
    else:
        raw_input = pd.read_csv(uploaded)

    raw, parse_warnings = standardize(raw_input)

    if parse_warnings:
        for w in parse_warnings:
            st.sidebar.warning(w)

    if "투자금액_백만원" in raw.columns and len(raw) > 0:
        result_df = run_all(raw)
        st.session_state["df"] = raw
        st.session_state["result_df"] = result_df
        st.session_state["summary"] = portfolio_summary(raw)
        st.sidebar.success(f"{len(raw)}개사 로드 완료")
    else:
        st.sidebar.error("필수 컬럼이 누락되었습니다. 입력 가이드를 참고하세요.")

elif use_sample:
    raw = load_portfolio("sample_portfolio.csv")
    result_df = run_all(raw)
    st.session_state["df"] = raw
    st.session_state["result_df"] = result_df
    st.session_state["summary"] = portfolio_summary(raw)
    st.sidebar.success("샘플 데이터(8개사) 로드됨")



# ── 탭 ───────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "Overview", "Fund Trend", "Analysis", "Benchmark", "Report",
])

# ── TAB 1: Performance ────────────────────────────
with tab1:
    st.markdown("""
<div style="background:#ffffff;border:1px solid #e5e5e5;border-radius:10px;padding:16px 20px;margin-bottom:20px;">
  <div style="font-size:13px;color:#1a1a1a;font-weight:600;margin-bottom:8px;">펀드 성과 한눈에 보기</div>
  <div style="display:flex;align-items:center;gap:6px;flex-wrap:wrap;font-size:11px;">
    <span style="background:#ffffff;color:#1a1a1a;border:1.5px solid #1b5e20;padding:3px 10px;border-radius:4px;font-weight:600;">1. 핵심 지표</span>
    <span style="color:#1b5e20;font-weight:700;">→</span>
    <span style="background:#ffffff;color:#1a1a1a;border:1.5px solid #1b5e20;padding:3px 10px;border-radius:4px;font-weight:600;">2. 포트폴리오 상세</span>
    <span style="color:#1b5e20;font-weight:700;">→</span>
    <span style="background:#ffffff;color:#1a1a1a;border:1.5px solid #1b5e20;padding:3px 10px;border-radius:4px;font-weight:600;">3. Top/Bottom</span>
    <span style="color:#1b5e20;font-weight:700;">→</span>
    <span style="background:#ffffff;color:#1a1a1a;border:1.5px solid #1b5e20;padding:3px 10px;border-radius:4px;font-weight:600;">4. 섹터·리스크</span>
  </div>
  <div style="font-size:10px;color:#999;margin-top:6px;">MOIC·IRR 등 핵심 성과부터 개별 기업, 섹터 집중도까지 펀드 전체를 파악합니다.</div>
</div>
""", unsafe_allow_html=True)
    if "result_df" not in st.session_state:
        st.info("사이드바에서 데이터를 로드하세요.")
    else:
        result_df = st.session_state["result_df"]
        df        = st.session_state["df"]
        summary   = st.session_state["summary"]
        _GREEN    = ["#1b5e20","#2e7d32","#388e3c","#43a047","#66bb6a","#81c784","#a5d6a7","#c8e6c9"]

        moic    = summary["펀드 MOIC"]
        tvpi    = summary["펀드 TVPI"]
        dpi     = summary["펀드 DPI"]
        rvpi    = summary["펀드 RVPI"]
        n       = summary["포트폴리오사 수"]
        avg_irr = round(result_df["IRR(%)"].mean(), 1)
        total_invested = df["투자금액_백만원"].sum()
        total_value    = df["현재가치_백만원"].sum() + df["회수금액_백만원"].sum()
        base_date      = pd.to_datetime(df["기준일"]).max().strftime("%Y-%m-%d")

        # ── 펀드 헤더 ───────────────────────────────
        st.markdown(f"""
<div style="display:flex; justify-content:space-between; align-items:baseline;
            margin-bottom:28px; flex-wrap:wrap; gap:8px;">
  <div>
    <div style="font-size:28px; font-weight:800; letter-spacing:-0.04em; color:#1a1a1a;">{fund_name}</div>
    <div style="font-size:12px; color:#bbb; margin-top:2px; letter-spacing:0.02em;">{fund_strategy} &middot; {quarter} &middot; {base_date}</div>
  </div>
  <div style="display:flex; gap:20px;">
    <div style="text-align:right;">
      <div style="font-size:10px; color:#bbb; letter-spacing:0.1em; text-transform:uppercase;">Portfolio</div>
      <div style="font-size:14px; font-weight:600; color:#1a1a1a;">{n}개사</div>
    </div>
    <div style="text-align:right;">
      <div style="font-size:10px; color:#bbb; letter-spacing:0.1em; text-transform:uppercase;">Invested</div>
      <div style="font-size:14px; font-weight:600; color:#1a1a1a;">{total_invested:,.0f}M</div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

        # ── Hero Metrics — MOIC + IRR ───────────────
        st.markdown(f"""
<div style="display:grid; grid-template-columns:1fr 1fr; gap:14px; margin-bottom:14px;">
  <div style="background:#ffffff; border:1px solid #c8e6c9; border-radius:12px; padding:28px 28px 24px;
              box-shadow: 0 4px 20px rgba(27,94,32,0.08);">
    <div class="stat-label">MOIC</div>
    <div class="stat-large" style="color:#1b5e20;">{moic}<span style="font-size:24px; color:#66bb6a;">x</span></div>
    <div class="stat-desc">투자원금 대비 전체 가치 배수</div>
  </div>
  <div style="background:#ffffff; border:1px solid #c8e6c9; border-radius:12px; padding:28px 28px 24px;
              box-shadow: 0 4px 20px rgba(27,94,32,0.08);">
    <div class="stat-label">IRR</div>
    <div class="stat-large" style="color:#1b5e20;">{avg_irr}<span style="font-size:24px; color:#66bb6a;">%</span></div>
    <div class="stat-desc">시간가치 반영 연환산 수익률</div>
  </div>
</div>
""", unsafe_allow_html=True)

        # ── 보조 Metrics — DPI · RVPI · TVPI ───────
        _dpi_status = "회수 진행" if dpi >= 0.5 else "초기 단계"
        _tvpi_color = "#1a1a1a"
        st.markdown(f"""
<div style="display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:12px;margin-bottom:14px;">
  <div style="background:#fff;border:1px solid #e5e5e5;border-radius:10px;padding:18px 20px;">
    <div style="font-size:10px;color:#999;font-weight:600;">DPI</div>
    <div style="font-size:28px;font-weight:800;color:#333;">{dpi}x</div>
    <div style="font-size:10px;color:#999;">{_dpi_status}</div>
  </div>
  <div style="background:#fff;border:1px solid #e5e5e5;border-radius:10px;padding:18px 20px;">
    <div style="font-size:10px;color:#999;font-weight:600;">RVPI</div>
    <div style="font-size:28px;font-weight:800;color:#333;">{rvpi}x</div>
    <div style="font-size:10px;color:#999;">잔존 가치</div>
  </div>
  <div style="background:#fff;border:1px solid {'#c8e6c9' if tvpi >= 2.0 else '#e5e5e5'};border-radius:10px;padding:18px 20px;">
    <div style="font-size:10px;color:#1a1a1a;font-weight:600;">TVPI</div>
    <div style="font-size:28px;font-weight:800;color:{_tvpi_color};">{tvpi}x</div>
    <div style="font-size:10px;color:#1a1a1a;">{'BM 달성' if tvpi >= 2.0 else 'DPI + RVPI'}</div>
  </div>
  <div style="background:#fff;border:1px solid #e5e5e5;border-radius:10px;padding:18px 20px;">
    <div style="font-size:10px;color:#999;font-weight:600;">COMPANIES</div>
    <div style="font-size:28px;font-weight:800;color:#333;">{n}개</div>
    <div style="font-size:10px;color:#999;">{int(total_invested):,}M 투자</div>
  </div>
</div>
""", unsafe_allow_html=True)

        # ── PE / VC 전략별 KPI ───────────────────────
        st.markdown("---")
        _is_vc = "VC" in fund_strategy
        _is_pe = "PE" in fund_strategy
        _is_growth = False

        if _is_vc:
            st.markdown("#### VC KPI — Growth Metrics")
            st.caption("VC 선택 시 표시되는 성장 중심 지표")
            vc_cols = ["회사명", "섹터", "투자단계", "투자금액_백만원", "현재가치_백만원", "MOIC", "IRR(%)"]
            vc_available = [c for c in vc_cols if c in result_df.columns]
            vc_df = result_df[vc_available].copy()
            # VC 관점: 성장률 기반 추정 (현재가치/투자금액 = 성장배수)
            vc_df["성장배수"] = (result_df["현재가치_백만원"] / result_df["투자금액_백만원"]).round(2)
            invest_days = (pd.to_datetime(df["기준일"]) - pd.to_datetime(df["투자일"])).dt.days
            vc_df["투자기간(월)"] = (invest_days / 30).astype(int).values
            vc_df["월평균성장률(%)"] = ((vc_df["성장배수"] ** (1 / (vc_df["투자기간(월)"].clip(lower=1) / 12)) - 1) * 100).round(1)

            kv1, kv2, kv3 = st.columns(3)
            kv1.metric("평균 성장배수", f"{vc_df['성장배수'].mean():.2f}x")
            kv2.metric("평균 투자기간", f"{vc_df['투자기간(월)'].mean():.0f}개월")
            kv3.metric("연환산 성장률", f"{vc_df['월평균성장률(%)'].mean():.1f}%")

            st.dataframe(
                vc_df[["회사명", "섹터", "투자단계", "성장배수", "투자기간(월)", "월평균성장률(%)", "MOIC"]],
                use_container_width=True, hide_index=True,
            )

        elif _is_pe:
            st.markdown("#### PE KPI — Value Metrics")
            st.caption("PE 선택 시 표시되는 밸류에이션 중심 지표")
            pe_df = result_df[["회사명", "섹터", "투자금액_백만원", "현재가치_백만원", "회수금액_백만원", "MOIC", "IRR(%)", "DPI"]].copy()
            pe_df["실현비율(%)"] = (result_df["회수금액_백만원"] / (result_df["현재가치_백만원"] + result_df["회수금액_백만원"]).clip(lower=1) * 100).round(1)
            pe_df["미실현가치"] = result_df["현재가치_백만원"].apply(lambda x: f"{int(x):,}")

            kp1, kp2, kp3 = st.columns(3)
            kp1.metric("평균 DPI", f"{dpi}x", help="현금 회수 배수")
            kp2.metric("평균 실현비율", f"{pe_df['실현비율(%)'].mean():.1f}%")
            kp3.metric("총 회수금액", f"{int(df['회수금액_백만원'].sum()):,}M")

            st.dataframe(
                pe_df[["회사명", "섹터", "MOIC", "IRR(%)", "DPI", "실현비율(%)", "미실현가치"]],
                use_container_width=True, hide_index=True,
            )

        else:
            st.markdown("#### Fund KPI")
            st.caption("펀드 종류에 따른 공통 지표")

        # ── Performance Summary ──────────────────────
        st.markdown("---")
        st.markdown("#### Performance")
        perf_data = {
            "지표": ["MOIC", "IRR", "DPI", "RVPI", "TVPI"],
            "값": [f"{moic}x", f"{avg_irr}%", f"{dpi}x", f"{rvpi}x", f"{tvpi}x"],
            "정의": [
                "투자원금 대비 전체 가치 배수",
                "현금흐름 시간가치 반영 연환산 수익률",
                "LP 출자금 대비 현금 회수 배수",
                "LP 출자금 대비 잔존 미실현 가치",
                "DPI + RVPI",
            ],
            "벤치마크": ["≥ 2.0x", "≥ 15%", "1.0x = 원금회수", "초기 높음", "≥ 2.0x"],
        }
        perf_df = pd.DataFrame(perf_data)
        st.dataframe(perf_df, use_container_width=True, hide_index=True, height=215)

        with st.expander("용어 해설"):
            st.markdown("""
| 지표 | 설명 |
|------|------|
| **MOIC** | 원금 대비 총 가치. 시간가치 미반영. 2x = 2배 |
| **IRR** | 투자·회수 타이밍 반영 연환산 수익률. 장기 보유 시 낮아짐 |
| **DPI** | 실제 현금 회수 배수. 1.0x 이상 = 원금 회수 완료 |
| **RVPI** | 미회수 평가가치 배수. 펀드 초기에 높음 |
| **TVPI** | DPI + RVPI. LP 관점 총 가치 배수 |
""")

        st.markdown("---")

        # ── Portfolio Detail ─────────────────────────
        st.markdown("#### Portfolio Detail")

        display_df = result_df[["회사명","섹터","투자단계","투자금액_백만원","MOIC","IRR(%)","DPI","RVPI","TVPI"]].copy()
        display_df = display_df.rename(columns={
            "투자금액_백만원": "투자금액(백만)",
            "IRR(%)": "IRR(%)",
        })
        display_df["투자금액(백만)"] = display_df["투자금액(백만)"].apply(lambda x: f"{int(x):,}")

        st.dataframe(display_df, use_container_width=True, hide_index=True, height=320)

        st.markdown("---")

        # ── Top / Bottom Performers ──────────────────
        st.markdown("#### Top / Bottom Performers")
        sorted_by_moic = result_df.sort_values("MOIC", ascending=False)
        top3 = sorted_by_moic.head(3)
        bottom3 = sorted_by_moic.tail(3)

        t_col, b_col = st.columns(2)
        with t_col:
            st.markdown('<p style="font-size:11px;letter-spacing:0.1em;color:#1b5e20;font-weight:700;">TOP PERFORMERS</p>', unsafe_allow_html=True)
            for _, r in top3.iterrows():
                irr_val = r["IRR(%)"]
                irr_color = "#1b5e20" if irr_val > 0 else "#c62828"
                st.markdown(f"""
<div style="background:#ffffff;border:1px solid #c8e6c9;border-radius:10px;padding:14px 18px;margin-bottom:8px;">
  <div style="display:flex;justify-content:space-between;align-items:center;">
    <div>
      <span style="font-size:15px;font-weight:700;color:#1a1a1a;">{r["회사명"]}</span>
      <span style="font-size:11px;color:#999;margin-left:8px;">{r["섹터"]}</span>
    </div>
    <div style="text-align:right;">
      <span style="font-size:20px;font-weight:800;color:#1b5e20;">{r["MOIC"]}x</span>
      <span style="font-size:12px;color:{irr_color};margin-left:8px;">IRR {irr_val}%</span>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

        with b_col:
            st.markdown('<p style="font-size:11px;letter-spacing:0.1em;color:#c62828;font-weight:700;">UNDERPERFORMERS</p>', unsafe_allow_html=True)
            for _, r in bottom3.iterrows():
                irr_val = r["IRR(%)"]
                moic_color = "#c62828" if r["MOIC"] < 1.0 else "#999"
                irr_color = "#c62828" if irr_val < 0 else "#999"
                st.markdown(f"""
<div style="background:#ffffff;border:1px solid #e5e5e5;border-radius:10px;padding:14px 18px;margin-bottom:8px;">
  <div style="display:flex;justify-content:space-between;align-items:center;">
    <div>
      <span style="font-size:15px;font-weight:700;color:#1a1a1a;">{r["회사명"]}</span>
      <span style="font-size:11px;color:#999;margin-left:8px;">{r["섹터"]}</span>
    </div>
    <div style="text-align:right;">
      <span style="font-size:20px;font-weight:800;color:{moic_color};">{r["MOIC"]}x</span>
      <span style="font-size:12px;color:{irr_color};margin-left:8px;">IRR {irr_val}%</span>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

        st.markdown("---")

        # ── 포트폴리오 집중도 (HHI) + 투자 경과 ────────
        st.markdown("#### Portfolio Analytics")

        an1, an2 = st.columns(2)
        with an1:
            st.markdown('<p style="font-size:11px;letter-spacing:0.1em;color:#999;font-weight:700;">CONCENTRATION (HHI)</p>', unsafe_allow_html=True)
            weights = df["투자금액_백만원"] / df["투자금액_백만원"].sum()
            hhi = round((weights ** 2).sum() * 10000)
            hhi_label = "높음 (집중)" if hhi > 2500 else ("보통" if hhi > 1500 else "낮음 (분산)")
            hhi_color = "#c62828" if hhi > 2500 else ("#ff9800" if hhi > 1500 else "#1b5e20")
            st.markdown(f"""
<div style="background:#ffffff;border:1px solid #e5e5e5;border-radius:10px;padding:20px 24px;">
  <div style="font-size:36px;font-weight:800;color:{hhi_color};">{hhi:,}</div>
  <div style="font-size:12px;color:#999;margin-top:4px;">HHI 지수 · {hhi_label}</div>
  <div style="margin-top:12px;background:#f0f0f0;border-radius:6px;height:8px;overflow:hidden;">
    <div style="width:{min(hhi/100, 100)}%;height:100%;background:{hhi_color};border-radius:6px;"></div>
  </div>
  <div style="display:flex;justify-content:space-between;margin-top:4px;">
    <span style="font-size:9px;color:#ccc;">0 (완전분산)</span>
    <span style="font-size:9px;color:#ccc;">10,000 (단일집중)</span>
  </div>
  <div style="font-size:10px;color:#aaa;margin-top:8px;line-height:1.5;">포트폴리오 내 투자금액 비중 기준. 특정 기업에 투자가 쏠릴수록 높아집니다.</div>
</div>
""", unsafe_allow_html=True)

        with an2:
            st.markdown('<p style="font-size:11px;letter-spacing:0.1em;color:#999;font-weight:700;">INVESTMENT TIMELINE</p>', unsafe_allow_html=True)
            avg_days = (pd.to_datetime(df["기준일"]) - pd.to_datetime(df["투자일"])).dt.days.mean()
            avg_months = int(avg_days / 30)
            avg_years = round(avg_days / 365.25, 1)
            realized_pct = round(df["회수금액_백만원"].sum() / (df["현재가치_백만원"].sum() + df["회수금액_백만원"].sum()) * 100, 1) if (df["현재가치_백만원"].sum() + df["회수금액_백만원"].sum()) > 0 else 0
            st.markdown(f"""
<div style="background:#ffffff;border:1px solid #e5e5e5;border-radius:10px;padding:20px 24px;">
  <div style="display:flex;gap:24px;">
    <div>
      <div style="font-size:36px;font-weight:800;color:#1a1a1a;">{avg_years}<span style="font-size:16px;color:#999;">년</span></div>
      <div style="font-size:12px;color:#999;">평균 투자 기간</div>
    </div>
    <div>
      <div style="font-size:36px;font-weight:800;color:#1b5e20;">{realized_pct}<span style="font-size:16px;color:#999;">%</span></div>
      <div style="font-size:12px;color:#999;">실현 비율</div>
    </div>
  </div>
  <div style="margin-top:12px;background:#f0f0f0;border-radius:6px;height:8px;overflow:hidden;">
    <div style="width:{realized_pct}%;height:100%;background:#1b5e20;border-radius:6px;"></div>
  </div>
  <div style="display:flex;justify-content:space-between;margin-top:4px;">
    <span style="font-size:9px;color:#ccc;">미실현</span>
    <span style="font-size:9px;color:#ccc;">전액 실현</span>
  </div>
</div>
""", unsafe_allow_html=True)

        st.markdown("---")

        # ── Charts ────────────────────────────────────
        _GP = ["#2e7d32","#43a047","#66bb6a","#81c784","#a5d6a7","#c8e6c9","#e8f5e9","#f1f8e9"]
        _CS = dict(plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                   font=dict(family="Pretendard, sans-serif", color="#1a1a1a", size=11))

        # MOIC 분포 (가로 풀와이드)
        sorted_moic = result_df.sort_values("MOIC", ascending=True)
        colors_moic = ["rgba(27,94,32,0.6)" if m >= 2 else "rgba(67,160,71,0.5)" if m >= 1 else "rgba(224,160,160,0.6)" for m in sorted_moic["MOIC"]]
        fig_bar = go.Figure(go.Bar(
            x=sorted_moic["MOIC"].tolist(), y=sorted_moic["회사명"].tolist(),
            orientation="h", marker_color=colors_moic, marker_line_width=0,
            text=[f"{m}x · {s}" for m, s in zip(sorted_moic["MOIC"], sorted_moic["섹터"])],
            textposition="outside", textfont=dict(size=10, color="#555"),
        ))
        fig_bar.add_vline(x=1.0, line_dash="dot", line_color="#ccc", annotation_text="BM 1.0x", annotation_font_size=9, annotation_font_color="#999")
        fig_bar.add_vline(x=2.0, line_dash="dot", line_color="#bbb", annotation_text="BM 2.0x", annotation_font_size=9, annotation_font_color="#999")
        fig_bar.update_layout(**_CS, height=300, margin=dict(t=50, b=20, l=90, r=80),
            title=dict(text="MOIC Distribution", font=dict(size=14), x=0.02, y=0.97),
            xaxis=dict(showgrid=True, gridcolor="#f0f0f0", zeroline=False),
            yaxis=dict(showgrid=False), bargap=0.35)
        st.plotly_chart(fig_bar, use_container_width=True)

        st.markdown("<div style='height:28px;'></div>", unsafe_allow_html=True)

        col_a, col_b = st.columns(2, gap="large")
        with col_a:
            # IRR 분포 바 차트
            sorted_irr = result_df.sort_values("IRR(%)", ascending=True)
            colors_irr = ["rgba(27,94,32,0.6)" if v >= 20 else "rgba(67,160,71,0.5)" if v >= 0 else "rgba(224,160,160,0.6)" for v in sorted_irr["IRR(%)"]]
            fig_irr = go.Figure(go.Bar(
                x=sorted_irr["IRR(%)"].tolist(), y=sorted_irr["회사명"].tolist(),
                orientation="h", marker_color=colors_irr, marker_line_width=0,
                text=[f"{v}%" for v in sorted_irr["IRR(%)"]],
                textposition="outside", textfont=dict(size=10, color="#555"),
            ))
            fig_irr.add_vline(x=15, line_dash="dot", line_color="#bbb", annotation_text="목표 15%", annotation_font_size=9, annotation_font_color="#999")
            fig_irr.update_layout(**_CS, height=340, margin=dict(t=55, b=20, l=95, r=55),
                title=dict(text="IRR Distribution", font=dict(size=14), x=0.02, y=0.97),
                xaxis=dict(showgrid=True, gridcolor="#f0f0f0", zeroline=False, title="IRR (%)"),
                yaxis=dict(showgrid=False), bargap=0.35)
            st.plotly_chart(fig_irr, use_container_width=True)

        with col_b:
            # 섹터 비중 도넛
            sector_df = df.groupby("섹터")["투자금액_백만원"].sum().sort_values(ascending=False).reset_index()
            fig_pie = go.Figure(go.Pie(
                labels=sector_df["섹터"].tolist(), values=sector_df["투자금액_백만원"].tolist(),
                marker=dict(colors=_GP, line=dict(color="#fff", width=2)),
                textinfo="label+percent", textfont=dict(size=11), hole=0.4,
            ))
            fig_pie.update_layout(**_CS, height=340, margin=dict(t=55, b=20, l=20, r=20),
                title=dict(text="Sector Allocation", font=dict(size=14), x=0.02, y=0.97),
                showlegend=False)
            st.plotly_chart(fig_pie, use_container_width=True)




# ── TAB 2: Portfolio ──────────────────────────────
with tab2:
    st.markdown("""
<div style="background:#ffffff;border:1px solid #e5e5e5;border-radius:10px;padding:16px 20px;margin-bottom:20px;">
  <div style="font-size:13px;color:#1a1a1a;font-weight:600;margin-bottom:8px;">펀드 추이 분석</div>
  <div style="display:flex;align-items:center;gap:6px;flex-wrap:wrap;font-size:11px;">
    <span style="background:#ffffff;color:#1a1a1a;border:1.5px solid #1b5e20;padding:3px 10px;border-radius:4px;font-weight:600;">1. J-Curve</span>
    <span style="color:#1b5e20;font-weight:700;">→</span>
    <span style="background:#ffffff;color:#1a1a1a;border:1.5px solid #1b5e20;padding:3px 10px;border-radius:4px;font-weight:600;">2. 분기별 추이</span>
  </div>
  <div style="font-size:10px;color:#999;margin-top:6px;">펀드 설정 이후 현금흐름 곡선과 분기별 성과 변화를 추적합니다.</div>
</div>
""", unsafe_allow_html=True)

    with st.expander("J-Curve란?"):
        st.markdown("""
사모펀드·VC 펀드는 초기에 투자 집행으로 누적 현금흐름이 마이너스(−)로 진입합니다.
이후 포트폴리오사 가치가 성장해 회수가 이뤄지면 플러스(+)로 전환되는데,
이 흐름이 알파벳 J자 형태를 그려 J-Curve라고 부릅니다.
""")

    # 포트폴리오 데이터에서 자동 J-Curve 생성
    cf_df = None
    if "df" in st.session_state and "jcurve_trend" not in st.session_state:
        _raw = st.session_state["df"]
        _rows = []
        for _, r in _raw.iterrows():
            _rows.append({"날짜": r["투자일"], "현금흐름_백만원": -float(r["투자금액_백만원"])})
            if float(r.get("회수금액_백만원", 0)) > 0:
                _rows.append({"날짜": r["기준일"], "현금흐름_백만원": float(r["회수금액_백만원"])})
        if _rows:
            cf_df = pd.DataFrame(_rows)

    cf_upload = st.file_uploader("현금흐름 CSV 직접 업로드 (선택)", type="csv", key="cf")
    if cf_upload:
        cf_df = pd.read_csv(cf_upload)
    elif cf_df is None:
        load_cf_sample = st.button("샘플 현금흐름 불러오기")
        if load_cf_sample:
            try:
                cf_df = pd.read_csv("sample_cashflows.csv")
                st.success("샘플 현금흐름 로드됨")
            except FileNotFoundError:
                st.error("sample_cashflows.csv 파일이 없습니다.")

    if cf_df is not None:
        trend = j_curve_data(cf_df)
        st.session_state["jcurve_trend"] = trend
        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=trend["날짜"], y=trend["누적현금흐름"],
            mode="lines+markers", name="누적 순현금흐름",
            line=dict(color="#2e7d32", width=3, shape="spline"),
            fill="tozeroy", fillcolor="rgba(200,230,201,0.3)",
            marker=dict(color="#1b5e20", size=8, line=dict(width=2, color="#ffffff")),
        ))
        fig.add_hline(y=0, line_dash="dot", line_color="#c8e6c9",
                      annotation_text="Break-even",
                      annotation_font_color="#81c784", annotation_font_size=10)
        fig.update_layout(
            height=350, margin=dict(t=30,b=20,l=20,r=20),
            plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            font=dict(family="Pretendard, sans-serif", color="#1a1a1a", size=12),
            xaxis_title="", yaxis_title="누적 순현금흐름 (백만원)",
        )
        fig.update_xaxes(showgrid=False, zeroline=False, tickfont_size=10)
        fig.update_yaxes(showgrid=True, gridcolor="#f0f0f0", zeroline=False, tickfont_size=10)
        st.plotly_chart(fig, use_container_width=True)

        # J-Curve 결과 요약
        _mn = trend["누적현금흐름"].min()
        _cr = trend["누적현금흐름"].iloc[-1]
        _be = trend[trend["누적현금흐름"] >= 0]
        _be_dt = str(_be["날짜"].iloc[0])[:10] if not _be.empty else "미도달"
        _mn_dt = str(trend.loc[trend["누적현금흐름"].idxmin(), "날짜"])[:10]

        _jc1, _jc2, _jc3 = st.columns(3)
        with _jc1:
            st.metric("최대 투자 시점", _mn_dt, delta=f"{_mn:,.0f}백만원")
        with _jc2:
            st.metric("현재 누적 현금흐름", f"{_cr:,.0f}백만원",
                      delta="회수 진행 중" if _cr > _mn else "투자 집중 구간")
        with _jc3:
            st.metric("손익분기 시점", _be_dt,
                      delta="도달" if not _be.empty else "미도달")

        if _cr < 0:
            _recovery = round((_cr - _mn) / abs(_mn) * 100, 1) if _mn != 0 else 0
            st.markdown(f"""
<div style="background:#fafafa;border-radius:8px;padding:12px 16px;margin-top:8px;font-size:12px;color:#555;line-height:1.7;">
  최저점({_mn_dt}) 대비 <b>{_recovery}%</b> 회복했으며, 현재 <b>{abs(_cr):,.0f}백만원</b> 미회수 상태입니다.
  펀드가 J-Curve 상승 구간에 있으며, 추가 회수가 이뤄지면 손익분기에 도달할 수 있습니다.
</div>""", unsafe_allow_html=True)
        else:
            st.markdown(f"""
<div style="background:#e8f5e9;border-radius:8px;padding:12px 16px;margin-top:8px;font-size:12px;color:#1b5e20;line-height:1.7;">
  펀드가 손익분기를 통과했습니다. 현재 누적 순현금흐름 <b>{_cr:,.0f}백만원</b>으로 수익 실현 단계에 진입했습니다.
</div>""", unsafe_allow_html=True)
    else:
        st.info("현금흐름 CSV를 업로드하거나 샘플을 불러오세요.")

    # ── 분기별 추이 섹션 ───────────────────────────
    st.markdown("---")
    st.markdown("### 분기별 펀드 지표 추이")
    quarters = load_quarters()
    if not quarters:
        st.info("저장된 분기 데이터가 없습니다.\n\n데이터 로드 후 사이드바에서 [현재 데이터 저장]을 눌러 분기를 누적하세요.")
    else:
        trend_df = load_trend()
        _LINE_COLORS = {"TVPI": "#1b5e20", "DPI": "#43a047", "RVPI": "#a5d6a7"}
        fig_q = go.Figure()
        for metric in ["TVPI", "DPI", "RVPI"]:
            fig_q.add_trace(go.Scatter(
                x=trend_df["quarter"], y=trend_df[metric],
                mode="lines+markers", name=metric,
                line=dict(color=_LINE_COLORS[metric], width=3, shape="spline"),
                marker=dict(color=_LINE_COLORS[metric], size=9,
                            line=dict(width=2, color="#ffffff")),
            ))
        fig_q.update_layout(
            height=350, margin=dict(t=30,b=20,l=20,r=20),
            xaxis_title="", yaxis_title="배수 (x)",
            plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            font=dict(family="Pretendard, sans-serif", color="#1a1a1a", size=12),
            legend=dict(orientation="h", y=-0.15, bgcolor="rgba(0,0,0,0)", font_size=11),
        )
        fig_q.update_xaxes(showgrid=False, zeroline=False, tickfont_size=10)
        fig_q.update_yaxes(showgrid=True, gridcolor="#f0f0f0", zeroline=False, tickfont_size=10)
        st.plotly_chart(fig_q, use_container_width=True)
        st.dataframe(trend_df, use_container_width=True, hide_index=True)

        # QoQ 변화 카드
        if len(trend_df) >= 2:
            _prev = trend_df.iloc[-2]
            _curr = trend_df.iloc[-1]
            st.markdown(f"##### QoQ 변화 ({_prev['quarter']} → {_curr['quarter']})")
            _qc1, _qc2, _qc3, _qc4 = st.columns(4)
            for _col, _metric in [(_qc1, "MOIC"), (_qc2, "TVPI"), (_qc3, "DPI"), (_qc4, "RVPI")]:
                if _metric in _curr and _metric in _prev:
                    _delta = round(_curr[_metric] - _prev[_metric], 2)
                    _col.metric(_metric, f"{_curr[_metric]}x", delta=f"{_delta:+.2f}x QoQ")

# ── TAB 3: Analysis ──────────────────────────────
with tab3:
    st.markdown("""
<div style="background:#ffffff;border:1px solid #e5e5e5;border-radius:10px;padding:16px 20px;margin-bottom:20px;">
  <div style="font-size:13px;color:#1a1a1a;font-weight:600;margin-bottom:8px;">분석 흐름</div>
  <div style="display:flex;align-items:center;gap:6px;flex-wrap:wrap;font-size:11px;">
    <span style="background:#ffffff;color:#1a1a1a;border:1.5px solid #1b5e20;padding:3px 10px;border-radius:4px;font-weight:600;">1. 기업 재무</span>
    <span style="color:#1b5e20;font-weight:700;">→</span>
    <span style="background:#ffffff;color:#1a1a1a;border:1.5px solid #1b5e20;padding:3px 10px;border-radius:4px;font-weight:600;">2. Exit 시나리오</span>
    <span style="color:#1b5e20;font-weight:700;">→</span>
    <span style="background:#ffffff;color:#1a1a1a;border:1.5px solid #1b5e20;padding:3px 10px;border-radius:4px;font-weight:600;">3. IRR Sensitivity</span>
    <span style="color:#1b5e20;font-weight:700;">→</span>
    <span style="background:#ffffff;color:#1a1a1a;border:1.5px solid #1b5e20;padding:3px 10px;border-radius:4px;font-weight:600;">4. Waterfall 분배</span>
  </div>
  <div style="font-size:10px;color:#999;margin-top:6px;">DART 재무 데이터로 기업을 파악한 후, Exit 시나리오와 수익 분배까지 이어지는 분석 흐름입니다.</div>
</div>
""", unsafe_allow_html=True)

    # ── ① DART 재무 조회 ────────────────────────────
    st.markdown("### 1. 기업 재무 조회")
    st.caption("DART 공시 재무제표를 조회하여 포트폴리오사의 매출·영업이익·순이익을 확인합니다.")
    corp_name = st.text_input("기업명 입력 (예: 삼성전자, 카카오)")

    if st.button("검색", key="dart_search") and corp_name:
        with st.spinner("DART에서 검색 중..."):
            results = search_company(corp_name)
        if results:
            st.session_state["dart_results"] = results
            st.session_state["dart_fin_df"] = None
        else:
            st.warning("검색 결과가 없습니다.")

    if "dart_results" in st.session_state and st.session_state["dart_results"]:
        options = {r["corp_name"]: r["corp_code"] for r in st.session_state["dart_results"]}
        selected = st.selectbox("기업 선택", list(options.keys()))

        @st.cache_data(ttl=3600, show_spinner=False)
        def _cached_financials(corp_code):
            return get_financials(corp_code)

        if st.button("재무제표 조회", key="dart_fin"):
            with st.spinner("재무제표 불러오는 중..."):
                fin_df = _cached_financials(options[selected])
            if not fin_df.empty:
                st.session_state["dart_fin_df"] = fin_df
                st.session_state["dart_selected"] = selected
            else:
                st.warning("재무 데이터를 불러올 수 없습니다.")

        if st.session_state.get("dart_fin_df") is not None:
            fin_df = st.session_state["dart_fin_df"]
            selected_name = st.session_state.get("dart_selected", selected)

            # 백만원 단위 표시용 DataFrame
            display_df = fin_df.copy()
            for col in ["매출액", "영업이익", "당기순이익"]:
                if col in display_df.columns:
                    display_df[col] = display_df[col].apply(
                        lambda x: f"{x/1e6:,.0f}" if pd.notna(x) and x != 0 else "-"
                    )
            display_df = display_df.rename(columns={"매출액": "매출액 (백만원)", "영업이익": "영업이익 (백만원)", "당기순이익": "당기순이익 (백만원)"})
            st.dataframe(display_df, use_container_width=True, hide_index=True)

            # 차트도 백만원 단위
            chart_df = fin_df.copy()
            for col in ["매출액", "영업이익", "당기순이익"]:
                if col in chart_df.columns:
                    chart_df[col] = chart_df[col].apply(lambda x: round(x / 1e6) if pd.notna(x) else 0)
            fig_dart = px.bar(
                chart_df.melt(id_vars="연도", value_vars=["매출액", "영업이익", "당기순이익"]),
                x="연도", y="value", color="variable", barmode="group",
                color_discrete_sequence=["#1b5e20", "#43a047", "#c8e6c9"],
                labels={"value": "금액 (백만원)", "variable": ""},
            )
            fig_dart.update_traces(marker_line_width=0, opacity=0.9)
            fig_dart.update_layout(
                height=320, margin=dict(t=20,b=20,l=20,r=20),
                plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                font=dict(family="Pretendard, sans-serif", color="#1a1a1a", size=12),
                legend=dict(orientation="h", y=-0.15, bgcolor="rgba(0,0,0,0)", font_size=10),
                bargap=0.5,
            )
            fig_dart.update_xaxes(showgrid=False, zeroline=False)
            fig_dart.update_yaxes(showgrid=True, gridcolor="#f0f0f0", zeroline=False, tickformat=",")
            st.plotly_chart(fig_dart, use_container_width=True)

            st.caption("Report 탭에서 'DART 재무분석'을 선택하면 보고서에 포함됩니다.")

    # ── ② 시나리오 시뮬레이터 ───────────────────────
    st.markdown("---")
    st.markdown("### 2. 회수 시나리오 시뮬레이터")
    st.caption("위 재무 데이터를 참고하여 포트폴리오사별 Exit 배수에 따른 IRR을 시뮬레이션합니다.")
    if "result_df" not in st.session_state:
        st.info("먼저 대시보드에서 데이터를 로드하세요.")
    else:
        result_df = st.session_state["result_df"]
        df        = st.session_state["df"]
        company2  = st.selectbox("포트폴리오사 선택", result_df["회사명"].tolist(), key="sim_company")
        r2        = result_df[result_df["회사명"] == company2].iloc[0]
        raw_r2    = df[df["회사명"] == company2].iloc[0]

        left2, right2 = st.columns([1, 2])
        with left2:
            st.metric("투자금액",      f"{int(raw_r2['투자금액_백만원']):,}백만원")
            st.metric("현재 MOIC",     f"{r2['MOIC']}x")
            st.metric("현재 IRR",      f"{r2['IRR(%)']}%")
            target_irr2 = st.slider("목표 IRR (%)", 10, 40, 20, key="sim_irr")
            opt2 = optimal_exit_timing(
                raw_r2["투자금액_백만원"], raw_r2["현재가치_백만원"],
                raw_r2["투자일"], target_irr2,
            )
            st.info(
                f"목표 IRR {target_irr2}% 달성 최소 배수: **{opt2[f'IRR {target_irr2}% 달성 최소 배수']}x**\n\n"
                f"{opt2['목표 달성 여부']}"
            )
        with right2:
            sim_df2 = simulate_exit(
                raw_r2["투자금액_백만원"], raw_r2["투자일"],
                [0.5, 1.0, 1.5, 2.0, 2.5, 3.0, 4.0, 5.0],
            )
            st.session_state["scenario_sim_df"] = sim_df2
            st.session_state["scenario_opt"] = opt2
            st.session_state["scenario_company"] = company2
            fig_sim2 = px.bar(
                sim_df2, x="Exit 배수", y="IRR (%)",
                color="IRR (%)", color_continuous_scale=["#e8f5e9","#66bb6a","#2e7d32","#1b5e20"],
                text="IRR (%)",
            )
            fig_sim2.add_hline(
                y=target_irr2, line_dash="dot", line_color="#a5d6a7",
                annotation_text=f"Target {target_irr2}%",
                annotation_font_color="#66bb6a", annotation_font_size=10,
            )
            fig_sim2.update_traces(texttemplate="%{text}%", textposition="outside",
                                   marker_line_width=0)
            fig_sim2.update_layout(
                height=320, margin=dict(t=20,b=20,l=20,r=20),
                plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                font=dict(family="Pretendard, sans-serif", color="#1a1a1a", size=12),
                bargap=0.3, showlegend=False,
            )
            fig_sim2.update_xaxes(showgrid=False, zeroline=False)
            fig_sim2.update_yaxes(showgrid=True, gridcolor="#f0f0f0", zeroline=False)
            st.plotly_chart(fig_sim2, use_container_width=True)
            st.dataframe(sim_df2, use_container_width=True, hide_index=True)


    # ── ③ IRR Sensitivity Matrix ────────────────────
    st.markdown("---")
    st.markdown("### 3. IRR Sensitivity Matrix")
    st.caption("시나리오 분석을 확장하여 Exit 배수 × 보유기간의 모든 조합에 대한 IRR을 한눈에 확인합니다.")
    st.caption("Exit 배수 × 보유기간 조합별 예상 IRR")

    if "df" not in st.session_state:
        st.info("대시보드에서 데이터를 먼저 로드하세요.")
    else:
        st.caption("개별 포트폴리오사 기준으로 Exit 배수와 보유기간에 따른 IRR을 분석합니다.")
        mat_company = st.selectbox(
            "분석 대상 기업", st.session_state["result_df"]["회사명"].tolist(), key="mat_co"
        )
        mat_raw = st.session_state["df"]
        mat_raw = mat_raw[mat_raw["회사명"] == mat_company].iloc[0]
        invested = float(mat_raw["투자금액_백만원"])

        multiples = [0.5, 0.75, 1.0, 1.25, 1.5, 2.0, 2.5, 3.0, 4.0, 5.0]
        years_list = list(range(1, 11))

        matrix = []
        for m in multiples:
            row = []
            for y in years_list:
                irr = round((m ** (1 / y) - 1) * 100, 1)
                row.append(irr)
            matrix.append(row)

        mat_df = pd.DataFrame(matrix, index=[f"{m}x" for m in multiples],
                              columns=[f"{y}년" for y in years_list])
        st.session_state["sensitivity_matrix_df"] = mat_df
        st.session_state["sensitivity_company"] = mat_company

        fig_mat = go.Figure(data=go.Heatmap(
            z=matrix,
            x=[f"{y}년" for y in years_list],
            y=[f"{m}x" for m in multiples],
            colorscale=[
                [0.0,  "#d32f2f"], [0.15, "#e53935"],
                [0.3,  "#ff9800"], [0.45, "#ffc107"],
                [0.55, "#cddc39"], [0.7,  "#66bb6a"],
                [0.85, "#43a047"], [1.0,  "#1b5e20"],
            ],
            zmid=15,
            text=[[f"{v}%" for v in row] for row in matrix],
            texttemplate="%{text}",
            textfont=dict(size=10, color="#ffffff"),
            hovertemplate="Exit %{y} · %{x}<br>IRR: %{z:.1f}%<extra></extra>",
        ))
        fig_mat.update_layout(
            xaxis_title="", yaxis_title="",
            height=400, margin=dict(t=20, b=20, l=20, r=20),
            plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            font=dict(family="Pretendard, sans-serif", color="#1a1a1a", size=12),
        )
        st.plotly_chart(fig_mat, use_container_width=True)
        st.caption("초록 = IRR 높음 / 빨강 = IRR 낮음 · 기준선 15% (일반적인 PE 목표 IRR)")

    # ── ④ Waterfall 계산기 ──────────────────────────
    st.markdown("---")
    st.markdown("### 4. Waterfall 분배 계산기")
    st.caption("위 시나리오에서 확인한 Exit 수익이 LP와 GP에게 어떻게 분배되는지 시뮬레이션합니다.")
    with st.expander("Waterfall이란?"):
        st.markdown("""
PE 펀드 회수금을 **LP → GP 순서로 단계별 분배**하는 구조입니다.
① 원금 반환 → ② Hurdle Rate 우선수익 → ③ GP 캐치업 → ④ 초과수익 분배 (Carried Interest).
GP는 Hurdle을 넘어야 Carry를 받을 수 있어 LP 이익 보호 장치로 작동합니다.
""")

    if "df" in st.session_state:
        _df_wf = st.session_state["df"]
        _auto_inv = int(_df_wf["투자금액_백만원"].sum())
        _auto_proc = int(_df_wf["현재가치_백만원"].sum() + _df_wf["회수금액_백만원"].sum())
    else:
        _auto_inv = 10000
        _auto_proc = 18000

    wf_c1, wf_c2, wf_c3 = st.columns(3)
    with wf_c1:
        wf_invested   = st.number_input("총 투자금액 (백만원)", min_value=100, value=_auto_inv, step=100, key="wf_inv")
        wf_proceeds   = st.number_input("총 회수금액 (백만원)", min_value=100, value=_auto_proc, step=100, key="wf_proc")
        wf_years      = st.number_input("펀드 기간 (년)", min_value=1, max_value=20, value=5, key="wf_years")
    with wf_c2:
        wf_hurdle     = st.slider("Hurdle Rate (%)", 0, 20, 8, key="wf_hurdle")
        wf_catchup    = st.slider("GP 캐치업율 (%)", 0, 100, 100, key="wf_catchup",
                                   help="100% = GP가 먼저 독식하며 catch-up / 50% = LP·GP 반반 분담")
    with wf_c3:
        wf_carry      = st.slider("Carried Interest (%)", 0, 30, 20, key="wf_carry")
        st.markdown("""
<div style="font-size:12px;color:#666;line-height:1.6;margin-top:8px;">
  <b>Hurdle Rate</b>: LP 우선수익률 (보통 8%)<br>
  <b>캐치업율</b>: GP가 Carry 몫을 따라잡는 속도<br>
  <b>Carried Interest</b>: GP 성과보수 비율 (보통 20%)
</div>
""", unsafe_allow_html=True)

    if st.button("Waterfall 계산", key="wf_calc", type="primary"):
        total_profit = max(0.0, float(wf_proceeds) - float(wf_invested))
        remaining    = float(wf_proceeds)
        steps        = []

        # ① 원금 반환
        lp1  = min(remaining, float(wf_invested))
        remaining -= lp1
        steps.append({"단계": "① 원금 반환", "LP": lp1, "GP": 0.0,
                       "누적 LP": lp1, "누적 GP": 0.0,
                       "설명": f"LP 투자원금 {wf_invested:,}백만원 전액 반환"})

        # ② 우선수익 (Hurdle)
        hurdle_amt = float(wf_invested) * ((1 + wf_hurdle/100) ** wf_years - 1)
        lp2  = min(remaining, hurdle_amt)
        remaining -= lp2
        steps.append({"단계": "② 우선수익 (Hurdle)", "LP": lp2, "GP": 0.0,
                       "누적 LP": lp1+lp2, "누적 GP": 0.0,
                       "설명": f"연 {wf_hurdle}% 복리 × {wf_years}년 → {hurdle_amt:,.0f}백만원 (실수취 {lp2:,.0f})"})

        # ③ GP 캐치업
        gp_carry_target = total_profit * wf_carry / 100
        if wf_catchup > 0 and remaining > 0 and gp_carry_target > 0:
            catchup_pool = min(remaining, gp_carry_target / (wf_catchup / 100))
            gp3  = catchup_pool * wf_catchup / 100
            lp3  = catchup_pool - gp3
            remaining -= catchup_pool
        else:
            gp3, lp3 = 0.0, 0.0
        steps.append({"단계": "③ GP 캐치업", "LP": lp3, "GP": gp3,
                       "누적 LP": lp1+lp2+lp3, "누적 GP": gp3,
                       "설명": f"GP Carry 목표 {gp_carry_target:,.0f}백만원 → {gp3:,.0f}백만원 확보 (캐치업율 {wf_catchup}%)"})

        # ④ 초과수익 분배
        gp4  = remaining * wf_carry / 100
        lp4  = remaining - gp4
        steps.append({"단계": "④ 초과수익 분배", "LP": lp4, "GP": gp4,
                       "누적 LP": lp1+lp2+lp3+lp4, "누적 GP": gp3+gp4,
                       "설명": f"LP {100-wf_carry}% ({lp4:,.0f}) / GP {wf_carry}% ({gp4:,.0f})"})

        total_lp = lp1+lp2+lp3+lp4
        total_gp = gp3+gp4
        eff_carry = total_gp/total_profit*100 if total_profit > 0 else 0

        # 요약 지표
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("총 투자금액", f"{wf_invested:,}백만원")
        m2.metric("총 회수금액", f"{wf_proceeds:,}백만원")
        m3.metric("LP 최종 수취", f"{total_lp:,.0f}백만원",
                  delta=f"MOIC {total_lp/wf_invested:.2f}x")
        m4.metric("GP Carry 수취", f"{total_gp:,.0f}백만원",
                  delta=f"실효 Carry {eff_carry:.1f}%")

        st.markdown("##### 단계별 분배 내역")
        step_df = pd.DataFrame(steps)[["단계","LP","GP","설명"]]
        step_df["LP"] = step_df["LP"].apply(lambda x: f"{x:,.0f}")
        step_df["GP"] = step_df["GP"].apply(lambda x: f"{x:,.0f}")
        st.dataframe(step_df, use_container_width=True, hide_index=True)

        # Waterfall 시각화 — LP vs GP 수평 바
        fig_wf = go.Figure()
        _wf_steps = ["① 원금 반환", "② 우선수익", "③ GP 캐치업", "④ 초과수익"]
        _wf_lp = [lp1, lp2, 0, lp4]
        _wf_gp = [0, 0, gp3, gp4]
        fig_wf.add_trace(go.Bar(name="LP", y=_wf_steps, x=_wf_lp, orientation="h",
            marker_color="rgba(27,94,32,0.7)", marker_line_width=0,
            text=[f"{v:,.0f}" if v > 0 else "" for v in _wf_lp], textposition="inside",
            textfont=dict(color="#fff", size=11)))
        fig_wf.add_trace(go.Bar(name="GP", y=_wf_steps, x=_wf_gp, orientation="h",
            marker_color="rgba(27,94,32,0.3)", marker_line_width=0,
            text=[f"{v:,.0f}" if v > 0 else "" for v in _wf_gp], textposition="inside",
            textfont=dict(color="#333", size=11)))
        fig_wf.update_layout(
            barmode="stack", height=280,
            plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            font=dict(family="Pretendard, sans-serif", color="#1a1a1a", size=11),
            margin=dict(t=10, b=10, l=100, r=20),
            xaxis=dict(showgrid=True, gridcolor="#f0f0f0", title="백만원"),
            yaxis=dict(showgrid=False, autorange="reversed"),
            legend=dict(orientation="h", y=-0.15, font_size=11),
            bargap=0.3,
        )
        st.plotly_chart(fig_wf, use_container_width=True)

        # LP vs GP 결과 요약
        _lp_moic = total_lp / wf_invested if wf_invested > 0 else 0
        _col_pie, _col_metrics = st.columns([1, 2])
        with _col_pie:
            fig_pie = go.Figure(go.Pie(
                values=[total_lp, total_gp], labels=["LP", "GP"],
                marker=dict(colors=["#1b5e20", "#c8e6c9"], line=dict(color="#fff", width=2)),
                textinfo="label+percent", textfont=dict(size=12),
                hole=0.45,
            ))
            fig_pie.update_layout(
                showlegend=False, height=220, margin=dict(t=10, b=10, l=10, r=10),
                paper_bgcolor="#ffffff",
            )
            st.plotly_chart(fig_pie, use_container_width=True)
        with _col_metrics:
            _r1, _r2 = st.columns(2)
            with _r1:
                st.metric("LP 수취", f"{total_lp:,.0f}M", delta=f"MOIC {_lp_moic:.2f}x")
                st.metric("LP 순수익", f"{total_lp-wf_invested:,.0f}M")
            with _r2:
                st.metric("GP Carry", f"{total_gp:,.0f}M", delta=f"실효 {eff_carry:.1f}%")
                st.metric("총 수익", f"{total_profit:,.0f}M")
        st.caption(f"Hurdle {wf_hurdle}% · 캐치업 {wf_catchup}% · Carry {wf_carry}% · {wf_years}년")

# ── TAB 4: Benchmark ─────────────────────────────
with tab4:
    st.markdown("""
<div style="background:#ffffff;border:1px solid #e5e5e5;border-radius:10px;padding:16px 20px;margin-bottom:20px;">
  <div style="font-size:13px;color:#1a1a1a;font-weight:600;margin-bottom:8px;">시장 환경 분석</div>
  <div style="display:flex;align-items:center;gap:6px;flex-wrap:wrap;font-size:11px;">
    <span style="background:#ffffff;color:#1a1a1a;border:1.5px solid #1b5e20;padding:3px 10px;border-radius:4px;font-weight:600;">1. KVIC 벤치마크</span>
    <span style="color:#1b5e20;font-weight:700;">→</span>
    <span style="background:#ffffff;color:#1a1a1a;border:1.5px solid #1b5e20;padding:3px 10px;border-radius:4px;font-weight:600;">2. 내 포트폴리오 vs 시장</span>
  </div>
  <div style="font-size:10px;color:#999;margin-top:6px;">국내 VC 시장 전체 동향과 내 펀드 성과를 비교합니다. 거시지표(금리·환율)는 보고서에 자동 포함됩니다.</div>
</div>
""", unsafe_allow_html=True)

    # 거시지표 데이터 백그라운드 로딩 (보고서용)
    @st.cache_data(ttl=3600, show_spinner=False)
    def _load_macro():
        r = get_base_rate(3)
        f = get_exchange_rate(3)
        return r, f

    rate_df, fx_df = _load_macro()
    st.session_state["macro_rate_df"] = rate_df
    st.session_state["macro_fx_df"] = fx_df
    if rate_df is not None and not rate_df.empty and "result_df" in st.session_state:
        latest_rate = rate_df["기준금리(%)"].iloc[-1]
        avg_irr = st.session_state["result_df"]["IRR(%)"].mean()
        st.session_state["macro_spread"] = round(avg_irr - latest_rate, 2)

    # ── KVIC 한국벤처투자 섹션
    st.markdown("---")
    st.markdown("### 한국벤처투자(KVIC) 모태펀드 벤치마크")
    st.caption("국내 VC 시장 전체 동향과 내 펀드 성과를 비교하기 위한 공공 데이터입니다.")

    import os
    if not os.getenv("KVIC_API_KEY"):
        st.info("KVIC_API_KEY가 없습니다. `.env`에 `KVIC_API_KEY=키값` 추가 후 재시작하세요.")
    else:
        col_l, col_r = st.columns(2)
        with col_l:
            kvic_year = st.selectbox(
                "조회 연도", list(range(2023, 2003, -1)), index=0, key="kvic_year"
            )
        with col_r:
            st.write("")
            load_kvic = st.button("KVIC 데이터 불러오기", key="kvic_load")

        if load_kvic:
            with st.spinner("한국벤처투자 API 조회 중..."):
                sector_df = get_kvic_sector_summary(kvic_year)
                trend_df_kvic = get_kvic_yearly_trend()
            st.session_state["kvic_sector"] = sector_df
            st.session_state["kvic_trend"] = trend_df_kvic

        if "kvic_sector" in st.session_state and not st.session_state["kvic_sector"].empty:
            sector_df = st.session_state["kvic_sector"]
            trend_df_kvic = st.session_state.get("kvic_trend", pd.DataFrame())

            total_funds = int(sector_df["조합수"].sum())
            total_amt = sector_df["총약정액(억원)"].sum()

            c1, c2 = st.columns(2)
            with c1:
                st.metric("전체 조합 수", f"{total_funds:,}개")
            with c2:
                st.metric("총 약정액", f"{total_amt:,.0f}억원")

            # 분야별 약정액 (상위 10)
            top10 = sector_df.head(10)
            fig_s = go.Figure(go.Bar(
                x=top10["총약정액(억원)"].tolist(),
                y=top10["투자분야"].tolist(),
                orientation="h", marker_color="rgba(27,94,32,0.6)", marker_line_width=0,
                text=[f"{v:,.0f}억 ({c}개)" for v, c in zip(top10["총약정액(억원)"], top10["조합수"])],
                textposition="outside", textfont=dict(size=10, color="#555"),
            ))
            fig_s.update_layout(
                title=dict(text=f"{kvic_year}년 모태펀드 분야별 약정액 (상위 10)", font=dict(size=14, color="#1a1a1a"), x=0.02),
                height=380, margin=dict(t=40, b=15, l=140, r=60),
                plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                font=dict(family="Pretendard, sans-serif", color="#1a1a1a", size=11),
                yaxis=dict(autorange="reversed", showgrid=False),
                xaxis=dict(showgrid=True, gridcolor="#f0f0f0", title="약정액 (억원)"),
                bargap=0.25,
            )
            st.plotly_chart(fig_s, use_container_width=True)

            # 연도별 결성 추이
            if not trend_df_kvic.empty:
                fig_t = go.Figure()
                fig_t.add_trace(go.Bar(
                    x=trend_df_kvic["결성연도"].tolist(),
                    y=trend_df_kvic["총약정액(억원)"].tolist(),
                    name="약정액", marker_color="rgba(27,94,32,0.5)", marker_line_width=0,
                    text=[f"{v:,.0f}억" for v in trend_df_kvic["총약정액(억원)"]],
                    textposition="outside", textfont=dict(size=9),
                ))
                fig_t.add_trace(go.Scatter(
                    x=trend_df_kvic["결성연도"].tolist(),
                    y=trend_df_kvic["결성조합수"].tolist(),
                    name="조합수", mode="lines+markers+text",
                    line=dict(color="#1b5e20", width=2),
                    marker=dict(size=6, color="#1b5e20"),
                    text=[f"{int(v)}개" for v in trend_df_kvic["결성조합수"]],
                    textposition="top center", textfont=dict(size=9, color="#1b5e20"),
                    yaxis="y2",
                ))
                fig_t.update_layout(
                    title=dict(text="연도별 모태펀드 결성 추이", font=dict(size=14, color="#1a1a1a"), x=0.02),
                    height=350, margin=dict(t=40, b=15, l=40, r=50),
                    plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                    font=dict(family="Pretendard, sans-serif", color="#1a1a1a", size=11),
                    legend=dict(orientation="h", y=-0.12, font_size=10),
                    xaxis=dict(showgrid=False),
                    yaxis=dict(showgrid=True, gridcolor="#f0f0f0", title="약정액 (억원)"),
                    yaxis2=dict(title="조합수", overlaying="y", side="right", showgrid=False),
                    bargap=0.3,
                )
                st.plotly_chart(fig_t, use_container_width=True)

                if not trend_df_kvic.empty and "result_df" in st.session_state:
                    my_inv = st.session_state["df"]["투자금액_백만원"].sum() / 100
                    fig_comp = go.Figure()
                    fig_comp.add_trace(go.Bar(
                        x=trend_df_kvic["결성연도"].tolist(), y=trend_df_kvic["총약정액(억원)"].tolist(),
                        name="KVIC 시장 (억원)", marker_color="#c8e6c9", marker_line_width=0,
                    ))
                    fig_comp.add_hline(y=my_inv, line_dash="dot", line_color="#1b5e20", line_width=2,
                                       annotation_text=f"내 펀드 {my_inv:,.0f}억", annotation_font_color="#1b5e20")
                    fig_comp.update_layout(
                        title=dict(text="시장 규모 대비 내 펀드 위치", font=dict(size=14)),
                        height=300, margin=dict(t=40, b=20, l=20, r=20),
                        plot_bgcolor="#ffffff", paper_bgcolor="#ffffff", font_color="#1a1a1a",
                        showlegend=True, legend=dict(orientation="h", y=-0.15),
                        yaxis_title="억원", bargap=0.3,
                    )
                    st.plotly_chart(fig_comp, use_container_width=True)

            st.divider()
            with st.expander("전체 분야별 데이터 보기"):
                st.dataframe(sector_df, use_container_width=True, hide_index=True)

    # ── 내 포트폴리오 vs KVIC 시장 비교 ──
    st.markdown("---")
    st.markdown("### 2. 내 포트폴리오 vs KVIC 시장")
    st.caption("KVIC 모태펀드 데이터 기반으로 국내 VC 시장에서 내 펀드가 어디에 위치하는지 분석합니다.")

    if "kvic_sector" not in st.session_state or st.session_state["kvic_sector"].empty:
        st.info("위에서 KVIC 데이터를 먼저 불러오세요.")
    elif "result_df" not in st.session_state:
        st.info("사이드바에서 포트폴리오 데이터를 먼저 로드하세요.")
    else:
        _my_df = st.session_state["result_df"]
        _my_raw = st.session_state["df"]
        _my_sum = st.session_state["summary"]
        _kvic_sec = st.session_state["kvic_sector"]
        _kvic_trend = st.session_state.get("kvic_trend", pd.DataFrame())

        kvic_total_amt = _kvic_sec["총약정액(억원)"].sum()
        kvic_total_funds = int(_kvic_sec["조합수"].sum())
        my_inv_억 = _my_raw["투자금액_백만원"].sum() / 100
        avg_per_fund = kvic_total_amt / kvic_total_funds if kvic_total_funds > 0 else 0
        share_pct = my_inv_억 / kvic_total_amt * 100 if kvic_total_amt > 0 else 0
        ratio = my_inv_억 / avg_per_fund if avg_per_fund > 0 else 0

        # ── 규모 비교 카드 (HTML) ──
        st.markdown(f"""
<div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;margin-bottom:20px;">
  <div style="background:#ffffff;border:1px solid #e5e5e5;border-radius:12px;padding:20px;text-align:center;">
    <div style="font-size:10px;color:#999;letter-spacing:0.08em;font-weight:600;margin-bottom:6px;">내 펀드 규모</div>
    <div style="font-size:28px;font-weight:800;color:#1b5e20;">{my_inv_억:,.0f}<span style="font-size:14px;font-weight:500;">억원</span></div>
    <div style="font-size:11px;color:#666;margin-top:4px;">KVIC 전체의 {share_pct:.2f}%</div>
  </div>
  <div style="background:#ffffff;border:1px solid #e5e5e5;border-radius:12px;padding:20px;text-align:center;">
    <div style="font-size:10px;color:#999;letter-spacing:0.08em;font-weight:600;margin-bottom:6px;">KVIC 시장 전체</div>
    <div style="font-size:28px;font-weight:800;color:#1a1a1a;">{kvic_total_amt:,.0f}<span style="font-size:14px;font-weight:500;">억원</span></div>
    <div style="font-size:11px;color:#666;margin-top:4px;">{kvic_total_funds:,}개 조합 운용 중</div>
  </div>
  <div style="background:#ffffff;border:1px solid #e5e5e5;border-radius:12px;padding:20px;text-align:center;">
    <div style="font-size:10px;color:#999;letter-spacing:0.08em;font-weight:600;margin-bottom:6px;">조합 평균 대비</div>
    <div style="font-size:28px;font-weight:800;color:{"#1b5e20" if ratio >= 1 else "#c62828"};">{ratio:.1f}<span style="font-size:14px;font-weight:500;">배</span></div>
    <div style="font-size:11px;color:#666;margin-top:4px;">평균 {avg_per_fund:,.0f}억원/조합</div>
  </div>
</div>
""", unsafe_allow_html=True)

        # ── 섹터 비교 차트 ──
        st.markdown("#### 섹터별 투자 비중 비교")
        st.caption("내 포트폴리오 섹터별 투자 규모와 KVIC 모태펀드 동일 분야 약정액을 비교합니다.")

        my_sectors = _my_df.groupby("섹터").agg(
            기업수=("회사명", "count"), 투자액=("투자금액_백만원", "sum")
        ).sort_values("투자액", ascending=False).reset_index()
        my_sectors["투자액_억"] = my_sectors["투자액"] / 100

        kvic_map = {}
        for sec in my_sectors["섹터"].tolist():
            for _, kr in _kvic_sec.iterrows():
                field = str(kr.get("투자분야", ""))
                if sec in field or field in sec:
                    kvic_map[sec] = {"약정액": kr["총약정액(억원)"], "조합수": int(kr["조합수"])}
                    break

        secs = my_sectors["섹터"].tolist()
        col_chart, col_detail = st.columns([3, 1])
        with col_chart:
            fig_compare = go.Figure()
            fig_compare.add_trace(go.Bar(
                x=secs, y=my_sectors["투자액_억"].tolist(),
                name="내 펀드", marker_color="#1b5e20", marker_line_width=0,
                text=[f"{v:,.0f}억" for v in my_sectors["투자액_억"]], textposition="outside",
                textfont=dict(size=11, color="#1b5e20"),
            ))
            if kvic_map:
                kvic_scaled = [kvic_map.get(s, {}).get("약정액", 0) / max(kvic_total_amt / my_inv_억, 1) for s in secs]
                fig_compare.add_trace(go.Bar(
                    x=secs, y=kvic_scaled,
                    name="KVIC (스케일 조정)", marker_color="#c8e6c9", marker_line_width=0,
                ))
            fig_compare.update_layout(
                barmode="group", height=320, margin=dict(t=10, b=20, l=20, r=20),
                plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                font=dict(family="Pretendard, sans-serif", color="#1a1a1a", size=12),
                legend=dict(orientation="h", y=-0.18, font_size=11),
                bargap=0.3, yaxis=dict(showgrid=True, gridcolor="#f0f0f0", title="억원"),
                xaxis=dict(showgrid=False),
            )
            st.plotly_chart(fig_compare, use_container_width=True)

        with col_detail:
            for sec in secs:
                my_amt = my_sectors[my_sectors["섹터"]==sec]["투자액_억"].values[0]
                kvic_info = kvic_map.get(sec)
                if kvic_info:
                    st.markdown(f"""
<div style="background:#fafafa;border-radius:8px;padding:8px 12px;margin-bottom:6px;">
  <div style="font-size:12px;font-weight:600;color:#1a1a1a;">{sec}</div>
  <div style="font-size:10px;color:#666;">내 펀드 {my_amt:,.0f}억 · KVIC {kvic_info['약정액']:,.0f}억 ({kvic_info['조합수']}개)</div>
</div>""", unsafe_allow_html=True)
                else:
                    st.markdown(f"""
<div style="background:#fafafa;border-radius:8px;padding:8px 12px;margin-bottom:6px;">
  <div style="font-size:12px;font-weight:600;color:#1a1a1a;">{sec}</div>
  <div style="font-size:10px;color:#999;">KVIC 매칭 분야 없음</div>
</div>""", unsafe_allow_html=True)

        # ── 연도별 시장 대비 위치 ──
        if not _kvic_trend.empty:
            st.markdown("#### 연간 결성 규모 대비 내 펀드 위치")
            st.caption("KVIC 모태펀드 연간 결성 규모와 내 펀드의 상대적 크기를 보여줍니다.")
            fig_pos = go.Figure()
            fig_pos.add_trace(go.Bar(
                x=_kvic_trend["결성연도"].tolist(),
                y=_kvic_trend["총약정액(억원)"].tolist(),
                name="KVIC 연간 결성", marker_color="#e8f5e9", marker_line_width=0,
                text=[f"{v:,.0f}" for v in _kvic_trend["총약정액(억원)"]], textposition="outside",
                textfont=dict(size=10),
            ))
            fig_pos.add_hline(y=my_inv_억, line_dash="dot", line_color="#1b5e20", line_width=2,
                              annotation_text=f"내 펀드 {my_inv_억:,.0f}억",
                              annotation_font=dict(color="#1b5e20", size=12))
            fig_pos.update_layout(
                height=300, margin=dict(t=10, b=20, l=20, r=20),
                plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                font=dict(color="#1a1a1a", size=12), showlegend=True,
                legend=dict(orientation="h", y=-0.18), bargap=0.3,
                yaxis=dict(showgrid=True, gridcolor="#f0f0f0", title="억원"),
            )
            st.plotly_chart(fig_pos, use_container_width=True)

        st.caption("KVIC 데이터: 한국벤처투자 공공 API 기준, 모태펀드 출자 조합 현황 반영")

# ── TAB 5: Report ────────────────────────────────
with tab5:
    st.markdown("### Report Builder")
    st.caption("포함할 섹션을 선택하고 LP 보고서(PDF) 또는 IC 장표(PPTX)를 생성하세요.")

    if "result_df" in st.session_state:
        result_df = st.session_state["result_df"]
        df        = st.session_state["df"]
        summary   = st.session_state["summary"]

        # multiselect 스타일
        st.markdown("""<style>
span[data-baseweb="tag"] { background-color:#ffffff !important; color:#1b5e20 !important; border:1.5px solid #1b5e20 !important;
    font-size:11px !important; height:26px !important; border-radius:6px !important; padding:0 8px !important; }
span[data-baseweb="tag"] span { color:#1a1a1a !important; }
span[data-baseweb="tag"] svg { fill:#999 !important; width:12px !important; }
</style>""", unsafe_allow_html=True)

        # 가이드 — 한 줄 안내
        st.markdown("""
<div style="display:flex;gap:24px;margin-bottom:12px;padding:10px 0;">
  <div style="display:flex;align-items:center;gap:6px;">
    <span style="background:#1b5e20;color:#fff;font-size:9px;font-weight:700;width:18px;height:18px;border-radius:50%;display:inline-flex;align-items:center;justify-content:center;">1</span>
    <span style="font-size:11px;color:#333;font-weight:600;">핵심 성과</span>
    <span style="font-size:10px;color:#aaa;">성과 · 포트폴리오 · 리스크</span>
  </div>
  <div style="display:flex;align-items:center;gap:6px;">
    <span style="background:#2e7d32;color:#fff;font-size:9px;font-weight:700;width:18px;height:18px;border-radius:50%;display:inline-flex;align-items:center;justify-content:center;">2</span>
    <span style="font-size:11px;color:#333;font-weight:600;">분석</span>
    <span style="font-size:10px;color:#aaa;">J-Curve · 시나리오 · Waterfall</span>
  </div>
  <div style="display:flex;align-items:center;gap:6px;">
    <span style="background:#43a047;color:#fff;font-size:9px;font-weight:700;width:18px;height:18px;border-radius:50%;display:inline-flex;align-items:center;justify-content:center;">3</span>
    <span style="font-size:11px;color:#333;font-weight:600;">부가</span>
    <span style="font-size:10px;color:#aaa;">거시지표 · KVIC · DART</span>
  </div>
</div>
""", unsafe_allow_html=True)

        _section_groups = [
            ("펀드 성과 종합", True, ["성과 요약", "포트폴리오 상세", "Top/Bottom", "섹터", "리스크", "집중도", "시각화 차트"]),
            ("J-Curve · 분기 추이", True, ["J-Curve", "분기별"]),
            ("시나리오 · Sensitivity", True, ["시나리오", "IRR Sensitivity"]),
            ("Waterfall 분배", True, ["Waterfall"]),
            ("KVIC 시장 비교", True, ["KVIC"]),
            ("거시지표 · DART", False, ["거시", "DART"]),
            ("AI 코멘터리", True, ["AI"]),
        ]
        st.markdown("##### 포함할 섹션")
        _cols = st.columns(4)
        selected = []
        for i, (label, default, keywords) in enumerate(_section_groups):
            with _cols[i % 4]:
                if st.checkbox(label, value=default, key=f"sec_{i}"):
                    selected.extend(keywords)
        st.session_state["report_sections"] = selected
        st.session_state["report_include_charts"] = "시각화 차트" in selected
        st.caption(f"{len(_section_groups)}개 중 {sum(1 for _,d,_ in _section_groups if d)}개 기본 선택. 아래 버튼을 누르면 생성됩니다.")

        st.markdown("---")

        # 출력 버튼
        if st.button("보고서 생성 (PPTX)", use_container_width=True):
            with st.spinner("PPTX 생성 중..."):
                from report_pptx import generate_lp_pptx
                _comm = generate_commentary(summary,
                    result_df[["회사명","MOIC","IRR(%)","TVPI","투자금액_백만원"]].to_dict("records")) if "AI" in str(selected) else ""
                _jc_pptx = st.session_state.get("jcurve_trend")
                if _jc_pptx is None and "J-Curve" in str(selected):
                    _jc_rows = []
                    for _, _r in df.iterrows():
                        _jc_rows.append({"날짜": _r["투자일"], "현금흐름_백만원": -float(_r["투자금액_백만원"])})
                        if float(_r.get("회수금액_백만원", 0)) > 0:
                            _jc_rows.append({"날짜": _r["기준일"], "현금흐름_백만원": float(_r["회수금액_백만원"])})
                    if _jc_rows:
                        from irr import j_curve_data
                        _jc_pptx = j_curve_data(pd.DataFrame(_jc_rows))
                _sc_pptx = st.session_state.get("scenario_sim_df")
                _sc_co_pptx = st.session_state.get("scenario_company", "펀드 전체")
                if _sc_pptx is None and "시나리오" in str(selected):
                    from simulator import simulate_exit
                    _sc_pptx = simulate_exit(float(df["투자금액_백만원"].sum()), df["투자일"].min(), [0.5,1.0,1.5,2.0,2.5,3.0,4.0,5.0])
                # Sensitivity 자동 생성
                _sens_pptx = st.session_state.get("sensitivity_matrix_df")
                _sens_co_pptx = st.session_state.get("sensitivity_company", "펀드 전체")
                if _sens_pptx is None and "Sensitivity" in str(selected):
                    _mults = [0.5, 1.0, 1.5, 2.0, 2.5, 3.0, 4.0, 5.0]
                    _yrs = list(range(1, 11))
                    _mat = [[round((m ** (1/y) - 1) * 100, 1) for y in _yrs] for m in _mults]
                    _sens_pptx = pd.DataFrame(_mat, index=[f"{m}x" for m in _mults], columns=[f"{y}년" for y in _yrs])
                pptx_bytes = generate_lp_pptx(summary, result_df, _comm, quarter,
                    fund_name=fund_name, fund_strategy=fund_strategy, base_date=base_date,
                    selected_sections=selected,
                    include_charts=st.session_state.get("report_include_charts", True),
                    jcurve_df=_jc_pptx, scenario_df=_sc_pptx, scenario_company=_sc_co_pptx,
                    sensitivity_df=_sens_pptx, sensitivity_company=_sens_co_pptx,
                    dart_fin_df=st.session_state.get("dart_fin_df"),
                    dart_company=st.session_state.get("dart_selected", ""),
                    kvic_sector_df=st.session_state.get("kvic_sector"),
                    rate_df=st.session_state.get("macro_rate_df"),
                    fx_df=st.session_state.get("macro_fx_df"),
                    spread=st.session_state.get("macro_spread"),
                    df_raw=df)
            st.download_button("PPTX 다운로드", pptx_bytes, file_name=f"Report_{quarter}.pptx",
                               mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                               use_container_width=True)

        # Excel 데이터 내보내기
        with st.expander("데이터 내보내기 (Excel)"):
            st.caption("분석 결과를 실무용 Excel 워크북으로 내보냅니다. PPT 선택 섹션과 자동 연동됩니다.")
            _full = result_df.copy()
            _full["투자기간(년)"] = ((pd.to_datetime(_full["기준일"]) - pd.to_datetime(_full["투자일"])).dt.days / 365.25).round(1)
            _full["투자비중(%)"] = (_full["투자금액_백만원"] / _full["투자금액_백만원"].sum() * 100).round(1)
            _full["총가치(백만)"] = (_full["현재가치_백만원"] + _full["회수금액_백만원"]).round(0)
            _full["수익금액(백만)"] = (_full["총가치(백만)"] - _full["투자금액_백만원"]).round(0)
            _full["수익률(%)"] = (_full["수익금액(백만)"] / _full["투자금액_백만원"] * 100).round(1)
            _full["실현비율(%)"] = (_full["회수금액_백만원"] / _full["총가치(백만)"] * 100).where(_full["총가치(백만)"] > 0, 0).round(1)
            _full["투자금액(백만)"] = _full["투자금액_백만원"].apply(lambda x: round(x))
            _full["현재가치(백만)"] = _full["현재가치_백만원"].apply(lambda x: round(x))
            _full["회수금액(백만)"] = _full["회수금액_백만원"].apply(lambda x: round(x))
            _full["BM달성"] = _full["MOIC"].apply(lambda x: "O" if x >= 2.0 else "X")
            _full["등급"] = _full["MOIC"].apply(lambda x: "상위" if x >= 3.0 else "양호" if x >= 2.0 else "보통" if x >= 1.0 else "부진")
            _export = _full[["회사명","섹터","투자단계","투자금액(백만)","현재가치(백만)","회수금액(백만)","총가치(백만)",
                             "MOIC","IRR(%)","DPI","RVPI","TVPI","투자기간(년)","투자비중(%)",
                             "수익금액(백만)","수익률(%)","실현비율(%)","BM달성","등급"]].copy()
            _sec_csv = result_df.groupby("섹터").agg(
                기업수=("회사명","count"), 총투자백만=("투자금액_백만원","sum"), 총현재가치백만=("현재가치_백만원","sum"),
                총회수백만=("회수금액_백만원","sum"), 평균MOIC=("MOIC","mean"), 평균IRR=("IRR(%)","mean"),
                최고MOIC=("MOIC","max"), 최저MOIC=("MOIC","min"),
            ).round(2).reset_index()
            _sec_csv["섹터비중(%)"] = (_sec_csv["총투자백만"] / _sec_csv["총투자백만"].sum() * 100).round(1)

            st.markdown("###### 전체 데이터 패키지 (Excel, 실무용)")
            st.caption("원본 데이터, 계산 지표+설명서, 섹터 요약 등을 하나의 워크북으로 제공합니다. 실사·감사 대응 시 활용하세요.")

            # ── 원본 데이터 ──
            _raw_cols = [c for c in ["회사명","섹터","투자단계","투자일","기준일",
                                      "투자금액_백만원","현재가치_백만원","회수금액_백만원"] if c in df.columns]
            _raw_export = df[_raw_cols].copy()

            # ── 계산지표 (해석 컬럼 추가) ──
            _export_rich = _export.copy()
            _export_rich["MOIC 해석"] = _export_rich["MOIC"].apply(
                lambda x: "최상위(3.0x+)" if x >= 3.0 else "우수(2.0x+)" if x >= 2.0 else "보통(1.0x+)" if x >= 1.0 else "부진(<1.0x)")
            _export_rich["IRR 해석"] = _export_rich["IRR(%)"].apply(
                lambda x: "최상위(25%+)" if x >= 25 else "우수(15%+)" if x >= 15 else "보통(10%+)" if x >= 10 else "저조(<10%)")
            _export_rich["DPI 해석"] = _export_rich["DPI"].apply(
                lambda x: "원금 전액 회수" if x >= 1.0 else "부분 회수" if x > 0 else "미회수")

            # ── 계산식 설명서 (합산용, 아래 블록으로 붙임) ──
            _weights = result_df["투자금액_백만원"] / result_df["투자금액_백만원"].sum()
            _hhi = round((_weights ** 2).sum() * 10000)
            _ex = result_df.iloc[0]
            _formula_dict = pd.DataFrame([
                {"지표": "MOIC", "분류": "성과지표",
                 "산식": "MOIC = (현재가치 + 회수금액) / 투자금액",
                 "계산 예시": f"{_ex['회사명']}: ({_ex['현재가치_백만원']:,.0f}+{_ex['회수금액_백만원']:,.0f}) / {_ex['투자금액_백만원']:,.0f} = {_ex['MOIC']}x",
                 "해석 기준": "1.0x 미만=원금손실 / 2.0x+=우수 / 3.0x+=최상위",
                 "활용 시점": "분기 성과 리뷰 필수 지표, LP 보고서 핵심 KPI"},
                {"지표": "DPI", "분류": "성과지표",
                 "산식": "DPI = 회수금액 / 투자금액",
                 "계산 예시": f"{_ex['회수금액_백만원']:,.0f} / {_ex['투자금액_백만원']:,.0f} = {_ex['DPI']}x",
                 "해석 기준": "0=미회수 / 1.0x=원금 전액 현금 회수 / 2.0x+=우수 Exit",
                 "활용 시점": "Exit 진행도 측정, 펀드 만기 근접 시 핵심 지표"},
                {"지표": "RVPI", "분류": "성과지표",
                 "산식": "RVPI = 현재가치(미실현) / 투자금액",
                 "계산 예시": f"{_ex['현재가치_백만원']:,.0f} / {_ex['투자금액_백만원']:,.0f} = {_ex['RVPI']}x",
                 "해석 기준": "펀드 초기 높음이 정상 / 후기까지 높으면 Exit 지연 의심",
                 "활용 시점": "미실현 포트폴리오 잔존가치 측정, 중간 보고"},
                {"지표": "TVPI", "분류": "성과지표",
                 "산식": "TVPI = DPI + RVPI = (회수금액 + 현재가치) / 투자금액",
                 "계산 예시": f"{_ex['DPI']} + {_ex['RVPI']} = {_ex['TVPI']}x",
                 "해석 기준": "MOIC와 동일 개념. LP 출자금(Paid-In) 기준 표기 시 사용",
                 "활용 시점": "LP 전체 수익 현황 파악, MOIC와 병행 표기"},
                {"지표": "IRR", "분류": "성과지표",
                 "산식": "Σ CFt/(1+r)^t = 0 을 만족하는 r. XIRR 방식으로 실제 날짜 기반 수치해석(scipy brentq)",
                 "계산 예시": f"투자 {_ex['투자금액_백만원']:,.0f}M({_ex['투자일']}) → 가치 {_ex['현재가치_백만원']+_ex['회수금액_백만원']:,.0f}M({_ex['기준일']}) → IRR={_ex['IRR(%)']}%",
                 "해석 기준": "10% 미만=저조 / 15%+=우수 / 25%+=최상위 (VC 목표: 20%+)",
                 "활용 시점": "펀드 수익률 BM 비교, PME 계산 입력값"},
                {"지표": "투자비중(%)", "분류": "포트폴리오",
                 "산식": "개별 투자금액 / 전체 투자금액 합계 × 100",
                 "계산 예시": "", "해석 기준": "단일 기업 20% 초과 시 과집중 경보",
                 "활용 시점": "자본 배분 리뷰, 포트폴리오 리밸런싱 판단"},
                {"지표": "실현비율(%)", "분류": "포트폴리오",
                 "산식": "회수금액 / (현재가치 + 회수금액) × 100",
                 "계산 예시": "", "해석 기준": "0%=전부 미실현 / 100%=전부 Exit 완료",
                 "활용 시점": "펀드 성숙도 측정, Exit 전략 수립"},
                {"지표": "HHI", "분류": "리스크",
                 "산식": "HHI = Σ(투자비중_i)² × 10,000",
                 "계산 예시": f"본 펀드 HHI = {_hhi:,} (투자비중 제곱합 × 10,000)",
                 "해석 기준": "1,500 미만=LOW / 1,500~2,500=MEDIUM / 2,500+=HIGH(집중 위험)",
                 "활용 시점": "포트폴리오 집중 리스크 정량화, LP 리스크 보고"},
                {"지표": "IRR Sensitivity", "분류": "시뮬레이션",
                 "산식": "IRR = (Exit 배수)^(1 / 보유기간) − 1",
                 "계산 예시": "Exit 2.0x, 보유 3년 → IRR = (2.0)^(1/3) − 1 = 26.0%",
                 "해석 기준": "행: Exit 배수(0.5x~5.0x) / 열: 보유기간(1~7년) 매트릭스",
                 "활용 시점": "시나리오 플래닝, Exit 타이밍·배수 목표 설정"},
                {"지표": "회수 시나리오", "분류": "시뮬레이션",
                 "산식": "회수금액 = 투자원금 × Exit 배수; IRR = XIRR(투자일~회수예정일)",
                 "계산 예시": "투자 1,000M × Exit 2.5x = 회수 2,500M → XIRR 계산",
                 "해석 기준": "Exit 배수 구간별 IRR 산출, 목표 IRR 달성 최소 배수 도출",
                 "활용 시점": "투심 자료, LP 배분 예상 커뮤니케이션"},
                {"지표": "목표 IRR 역산", "분류": "시뮬레이션",
                 "산식": "필요배수 = (1 + 목표IRR%)^경과연수",
                 "계산 예시": "목표 IRR 20%, 3년 경과 → (1.20)^3 = 1.728x → 최소 MOIC 1.73x 필요",
                 "해석 기준": "현재 MOIC ≥ 필요배수이면 현 시점 Exit으로 목표 IRR 달성 가능. 미달 시 추가 보유 또는 가치 제고 필요",
                 "활용 시점": "Exit 타이밍 판단, '지금 팔면 목표 수익 나오나?' 역산. simulate_exit()·optimal_exit_timing() 공용 함수 사용"},
                {"지표": "Waterfall ①원금반환", "분류": "분배구조",
                 "산식": "LP 원금 = MIN(회수재원, 투자원금)",
                 "계산 예시": "", "해석 기준": "LP 투자 원금 최우선 반환. GP는 이 단계에서 0원 배분",
                 "활용 시점": "실제 Exit 시 LP/GP 배분 계획 수립"},
                {"지표": "Waterfall ②우선수익", "분류": "분배구조",
                 "산식": "우선수익 = 투자원금 × [(1 + Hurdle%)^연수 − 1]",
                 "계산 예시": "9,050M × [(1.08)^5 − 1] = 4,247M  (Hurdle 8%, 5년)",
                 "해석 기준": "LP에게 약정 Hurdle Rate(통상 8%) 수익 우선 배분",
                 "활용 시점": "Carried Interest 계산 전 LP 보호 장치"},
                {"지표": "Waterfall ③GP Catch-up", "분류": "분배구조",
                 "산식": "GP Catch-up = MIN(잔여재원, 총수익 × Carry%)",
                 "계산 예시": "총수익의 20%를 GP가 우선 수취",
                 "해석 기준": "GP가 전체 수익 대비 약정 Carry 비율만큼 보전",
                 "활용 시점": "LP 우선수익 지급 직후 단계"},
                {"지표": "Waterfall ④Carry Split", "분류": "분배구조",
                 "산식": "잔여재원 × LP(1−Carry%) + GP(Carry%) 비율 분할",
                 "계산 예시": "잔여 × LP 80% / GP 20%",
                 "해석 기준": "①~③ 이후 초과 수익을 LP·GP가 약정 비율로 최종 분배",
                 "활용 시점": "최종 청산 시, 혹은 분기별 Escrow 계산"},
            ])

            # ── 펀드 요약 (항목 설명 추가) ──
            _metric_desc = {
                "포트폴리오사 수":       ("투자 집행 기업 수",                    "15~25개가 일반적인 적정 분산 범위"),
                "총 투자금액 (백만원)":  ("Paid-in Capital 합산",                 "실제 집행된 출자 원금"),
                "현재가치 합계 (백만원)":("미실현 NAV 합산",                     "매 분기 재평가 필요"),
                "회수금액 합계 (백만원)":("실현 Exit 금액 합산",                  ""),
                "펀드 MOIC":            ("총가치 / 투자금액 배수",                "2.0x=우수, 3.0x+=최상위"),
                "펀드 DPI":             ("현금 회수 배수",                        "1.0x=원금 전액 현금 회수"),
                "펀드 RVPI":            ("미실현 가치 배수",                      "초기 높음이 정상, 후기까지 높으면 Exit 지연"),
                "펀드 TVPI":            ("DPI + RVPI 합산 배수 (MOIC와 동일)",    ""),
            }
            _summary_sheet = pd.DataFrame([
                {"항목": k, "값": v,
                 "항목 설명": _metric_desc.get(k, ("",""))[0],
                 "해석 기준": _metric_desc.get(k, ("",""))[1]}
                for k, v in summary.items()
            ])

            # ── 섹터 요약 (순위·HHI기여도 추가) ──
            _sec_csv["투자비중순위"] = _sec_csv["총투자백만"].rank(ascending=False).astype(int)
            _sec_csv["HHI기여도"] = ((_sec_csv["총투자백만"] / _sec_csv["총투자백만"].sum()) ** 2 * 10000).round(0).astype(int)
            _sec_csv = _sec_csv[["투자비중순위","섹터","기업수","총투자백만","섹터비중(%)",
                                  "HHI기여도","평균MOIC","최고MOIC","최저MOIC","평균IRR",
                                  "총현재가치백만","총회수백만"]].sort_values("투자비중순위")

            # ── Waterfall (비중·누적 컬럼 추가) ──
            _wf_inv = float(total_invested)
            _wf_proc = float(result_df["현재가치_백만원"].sum() + result_df["회수금액_백만원"].sum())
            _hurdle, _carry, _years = 8, 20, 5
            _profit = max(0, _wf_proc - _wf_inv)
            _hurdle_amt = _wf_inv * ((1 + _hurdle / 100) ** _years - 1)
            _rem = _wf_proc
            _s1 = min(_rem, _wf_inv); _rem -= _s1
            _s2 = min(_rem, _hurdle_amt); _rem -= _s2
            _gp_t = _profit * _carry / 100
            _s3_gp = min(_rem, _gp_t); _rem -= _s3_gp
            _s4_gp = _rem * _carry / 100; _s4_lp = _rem - _s4_gp
            _total_lp = _s1 + _s2 + _s4_lp; _total_gp = _s3_gp + _s4_gp
            _cum_lp, _cum_gp = 0.0, 0.0
            _wf_data = []
            for _st, _lp, _gp, _desc in [
                ("① 원금반환",    _s1,    0,       f"LP 투자 원금 최우선 반환"),
                ("② 우선수익",    _s2,    0,       f"LP Hurdle {_hurdle}% × {_years}년 우선 배분"),
                ("③ GP Catch-up", 0,      _s3_gp,  "GP Carry 목표 보전 단계"),
                ("④ Carry Split", _s4_lp, _s4_gp,  f"LP {100-_carry}% / GP {_carry}% 최종 분배"),
            ]:
                _cum_lp += _lp; _cum_gp += _gp
                _wf_data.append({
                    "단계": _st, "LP(백만원)": round(_lp), "GP(백만원)": round(_gp),
                    "합계(백만원)": round(_lp + _gp),
                    "비중(%)": round((_lp + _gp) / _wf_proc * 100, 1) if _wf_proc > 0 else 0,
                    "누적 LP(백만원)": round(_cum_lp),
                    "누적 GP(백만원)": round(_cum_gp),
                    "단계 설명": _desc,
                })
            _wf_data.append({
                "단계": "합계", "LP(백만원)": round(_total_lp), "GP(백만원)": round(_total_gp),
                "합계(백만원)": round(_wf_proc), "비중(%)": 100.0,
                "누적 LP(백만원)": round(_total_lp), "누적 GP(백만원)": round(_total_gp),
                "단계 설명": (f"LP MOIC {_total_lp/_wf_inv:.2f}x / GP Carry {_total_gp:,.0f}백만원"
                              if _wf_inv > 0 else ""),
            })
            _wf_sheet = pd.DataFrame(_wf_data)

            # ── PPT 선택 섹션 → Excel 시트 자동 연동 ──
            _ppt_sel = st.session_state.get("report_sections", [])
            def _xl_inc(*keywords):
                """PPT 섹션 미선택 시 전부 포함, 선택된 경우엔 해당 키워드 일치 시만 포함"""
                if not _ppt_sel:
                    return True
                return any(any(kw in s for kw in keywords) for s in _ppt_sel)

            # 시트 포함 여부 계산
            _sheet_plan = [
                ("0_펀드요약",         True,                                "항상 포함"),
                ("1_원본데이터(RAW)",  True,                                "항상 포함"),
                ("2_계산지표&설명서",  True,                                "항상 포함"),
                ("3_섹터요약",         _xl_inc("섹터"),                     "섹터 선택 시"),
                ("4_J-Curve",          _xl_inc("J-Curve"),                  "J-Curve 선택 시"),
                ("5_IRR_Sensitivity",  _xl_inc("IRR Sensitivity"),          "시나리오·Sensitivity 선택 시"),
                ("6_회수시나리오",     _xl_inc("시나리오"),                  "시나리오 선택 시"),
                ("7_Waterfall분배",    _xl_inc("Waterfall"),                "Waterfall 선택 시"),
                ("8_KVIC벤치마크",     _xl_inc("KVIC"),                     "KVIC 선택 시"),
                ("9_거시지표",         _xl_inc("거시", "DART"),             "거시지표·DART 선택 시"),
            ]
            _inc_sheets = [n for n, ok, _ in _sheet_plan if ok]
            _exc_sheets = [(n, hint) for n, ok, hint in _sheet_plan if not ok]

            # 포함될 시트 미리보기 UI
            if _ppt_sel:
                _badge_html = " ".join(
                    f'<span style="background:#1b5e20;color:#fff;font-size:10px;padding:2px 8px;border-radius:10px;margin:2px">{n}</span>'
                    for n in _inc_sheets
                )
                if _exc_sheets:
                    _badge_html += " " + " ".join(
                        f'<span style="background:#eee;color:#999;font-size:10px;padding:2px 8px;border-radius:10px;margin:2px;text-decoration:line-through">{n}</span>'
                        for n, _ in _exc_sheets
                    )
                st.markdown(f"**PPT 선택 섹션 연동 — 포함 시트 {len(_inc_sheets)}개:**", unsafe_allow_html=False)
                st.markdown(_badge_html, unsafe_allow_html=True)
                if _exc_sheets:
                    st.caption("취소선 시트는 PPT에서 해당 섹션이 미선택되어 제외됩니다.")
            else:
                st.caption("PPT 섹션을 선택하면 Excel 시트가 자동으로 연동됩니다. (현재: 전체 포함)")

            # ── Excel 워크북 생성 ──
            _buf = io.BytesIO()
            # 계산지표+설명서 통합: 설명서 블록은 데이터 아래 3행 공백 후 시작
            _formula_start_row = len(_export_rich) + 3   # 0-based → Excel row = +1
            _formula_hdr_row   = _formula_start_row + 1  # 1-based Excel row of formula header

            with pd.ExcelWriter(_buf, engine="openpyxl") as _writer:
                _summary_sheet.to_excel(_writer, sheet_name="0_펀드요약", index=False)
                _raw_export.to_excel(_writer, sheet_name="1_원본데이터(RAW)", index=False)
                # 계산지표 + 설명서 통합 시트 (항상 포함)
                _export_rich.to_excel(_writer, sheet_name="2_계산지표&설명서", index=False, startrow=0)
                _formula_dict.to_excel(_writer, sheet_name="2_계산지표&설명서", index=False,
                                       startrow=_formula_start_row)
                if _xl_inc("섹터"):
                    _sec_csv.to_excel(_writer, sheet_name="3_섹터요약", index=False)
                if _xl_inc("J-Curve") and st.session_state.get("jcurve_trend") is not None:
                    st.session_state["jcurve_trend"].to_excel(_writer, sheet_name="4_J-Curve", index=False)
                if _xl_inc("IRR Sensitivity") and st.session_state.get("sensitivity_matrix_df") is not None:
                    st.session_state["sensitivity_matrix_df"].to_excel(_writer, sheet_name="5_IRR_Sensitivity")
                if _xl_inc("시나리오") and st.session_state.get("scenario_sim_df") is not None:
                    st.session_state["scenario_sim_df"].to_excel(_writer, sheet_name="6_회수시나리오", index=False)
                if _xl_inc("Waterfall"):
                    _wf_sheet.to_excel(_writer, sheet_name="7_Waterfall분배", index=False)
                if _xl_inc("KVIC") and st.session_state.get("kvic_sector") is not None and not st.session_state["kvic_sector"].empty:
                    st.session_state["kvic_sector"].to_excel(_writer, sheet_name="8_KVIC벤치마크", index=False)
                _macro_rows = []
                if _xl_inc("거시", "DART"):
                    if st.session_state.get("macro_rate_df") is not None and not st.session_state["macro_rate_df"].empty:
                        _macro_rows.append(("기준금리(%)", st.session_state["macro_rate_df"]["기준금리(%)"].iloc[-1]))
                    if st.session_state.get("macro_fx_df") is not None and not st.session_state["macro_fx_df"].empty:
                        _macro_rows.append(("원/달러(원)", st.session_state["macro_fx_df"]["원/달러(원)"].iloc[-1]))
                    if st.session_state.get("macro_spread") is not None:
                        _macro_rows.append(("펀드 스프레드(%p)", st.session_state["macro_spread"]))
                if _macro_rows:
                    pd.DataFrame(_macro_rows, columns=["항목", "값"]).to_excel(
                        _writer, sheet_name="9_거시지표", index=False)

                # ── 서식 적용 ──
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.formatting.rule import ColorScaleRule
                from openpyxl.utils import get_column_letter

                _H_FILL  = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
                _H_FONT  = Font(color="FFFFFF", bold=True, size=10)
                _ALT_FILL = PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid")
                _THIN = Side(style="thin", color="CCCCCC")
                _BDR  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
                _LONG_COLS = {"산식","계산 예시","해석 기준","활용 시점","단계 설명",
                              "항목 설명","설명","해석기준"}
                _NUM_FMT = '#,##0.##'

                for _sname, _ws in _writer.sheets.items():
                    _is_merged = _sname.startswith("2_")
                    # 헤더 행 (row 1)
                    _ws.row_dimensions[1].height = 30
                    for _cell in _ws[1]:
                        if _cell.value is not None:
                            _cell.fill = _H_FILL; _cell.font = _H_FONT
                            _cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                            _cell.border = _BDR

                    # 데이터 행
                    for _ridx in range(2, _ws.max_row + 1):
                        # 통합 시트: 설명서 헤더 행 처리
                        if _is_merged and _ridx == _formula_hdr_row:
                            _ws.row_dimensions[_ridx].height = 30
                            for _cell in _ws[_ridx]:
                                if _cell.value is not None:
                                    _cell.fill = _H_FILL; _cell.font = _H_FONT
                                    _cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                                    _cell.border = _BDR
                            continue

                        _fill = _ALT_FILL if _ridx % 2 == 0 else PatternFill(fill_type=None)
                        _ws.row_dimensions[_ridx].height = 18
                        for _cell in _ws[_ridx]:
                            _cell.border = _BDR
                            _cell.fill = _fill
                            if _cell.value is None:
                                continue
                            # 어느 헤더 행을 참조할지 결정 (통합 시트 하단 블록은 설명서 헤더 참조)
                            _ref_hdr = (_formula_hdr_row
                                        if (_is_merged and _ridx > _formula_hdr_row)
                                        else 1)
                            _col_hdr = _ws.cell(row=_ref_hdr, column=_cell.column).value or ""
                            if _col_hdr in _LONG_COLS:
                                _cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                                # 한글은 Excel에서 ASCII의 약 1.8배 폭 → 유효 길이로 줄 수 추정
                                _cval = str(_cell.value) if _cell.value else ""
                                _eff = sum(1.8 if ord(c) > 127 else 1.0 for c in _cval)
                                _lines = int(_eff // 40) + 1  # 열폭 65 기준 한글 혼합 약 40유닛/줄
                                _est_h = min(8 + _lines * 14, 90)  # 최대 90pt (약 6줄)
                                _ws.row_dimensions[_ridx].height = max(_ws.row_dimensions[_ridx].height, _est_h)
                            elif isinstance(_cell.value, (int, float)):
                                _cell.number_format = _NUM_FMT
                                _cell.alignment = Alignment(horizontal="right", vertical="center")
                            else:
                                _cell.alignment = Alignment(horizontal="left", vertical="center")

                    # 틀고정 + 자동필터 (첫 행 기준)
                    _ws.freeze_panes = "A2"
                    _hdr_range = f"A1:{get_column_letter(_ws.max_column)}1"
                    _ws.auto_filter.ref = _hdr_range

                    # 열 너비: 긴 텍스트 컬럼은 고정 폭, 나머지는 내용 기반 자동
                    for _col in _ws.columns:
                        _letter = _col[0].column_letter
                        # 통합 시트는 수식 설명서 헤더(하단 블록)도 같이 확인
                        _hdr_val = str(_ws[f"{_letter}1"].value or "")
                        _hdr_val2 = str(_ws.cell(row=_formula_hdr_row, column=_col[0].column).value or "") if _is_merged else ""
                        if _hdr_val in _LONG_COLS or _hdr_val2 in _LONG_COLS:
                            _ws.column_dimensions[_letter].width = 65
                        else:
                            # Korean 글자는 Excel에서 ASCII의 약 2배 폭 → 한글 비율 반영
                            def _ew(v):
                                s = str(v) if v is not None else ""
                                return sum(2 if ord(c) > 127 else 1 for c in s)
                            _max_len = max((_ew(_c.value) for _c in _col if _c.value is not None), default=8)
                            _ws.column_dimensions[_letter].width = min(max(_max_len + 4, 14), 55)

                # ── 조건부 서식: MOIC / IRR / DPI (빨강→노랑→초록) ──
                _ws2 = _writer.sheets["2_계산지표&설명서"]
                _cols2 = list(_export_rich.columns)
                for _mc, _c0, _c1, _c2 in [
                    ("MOIC",   "F4C7C3", "FFEB9C", "C6E0B4"),
                    ("IRR(%)", "F4C7C3", "FFEB9C", "C6E0B4"),
                    ("DPI",    "F4C7C3", "FFEB9C", "C6E0B4"),
                ]:
                    if _mc in _cols2:
                        _ci  = _cols2.index(_mc) + 1
                        _cl  = get_column_letter(_ci)
                        _rng = f"{_cl}2:{_cl}{len(_export_rich)+1}"
                        _ws2.conditional_formatting.add(_rng, ColorScaleRule(
                            start_type="min",        start_color=_c0,
                            mid_type="percentile",   mid_value=50, mid_color=_c1,
                            end_type="max",          end_color=_c2,
                        ))

            st.download_button("Excel 패키지 다운로드 (.xlsx)", _buf.getvalue(),
                               file_name=f"PortfolioDataPackage_{quarter}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
            _n_inc = len(_inc_sheets)
            st.caption(f"틀고정·필터·조건부서식(MOIC/IRR/DPI) 적용. {_n_inc}개 시트 — PPT 선택 섹션 연동{'됨' if _ppt_sel else ' (전체 포함)'}.")

        st.markdown("---")
        st.caption("본 보고서는 참고용 자료이며, 투자 결정의 근거로 단독 사용할 수 없습니다.")

    st.markdown("---")
    st.markdown("### AI 분석")
    if "result_df" not in st.session_state:
        st.info("먼저 대시보드에서 데이터를 로드하세요.")
    else:
        result_df = st.session_state["result_df"]
        summary   = st.session_state["summary"]

        col1, col2 = st.columns(2, gap="large")
        with col1:
            st.markdown("#### 분기 코멘터리")
            st.caption("Claude AI가 펀드 성과 데이터를 분석하여 LP 보고서용 코멘터리를 자동 작성합니다.")
            if st.button("코멘터리 생성"):
                with st.spinner("Claude가 작성 중..."):
                    detail_rows = result_df[["회사명", "MOIC", "IRR(%)", "TVPI", "투자금액_백만원"]].to_dict("records")
                    commentary = generate_commentary(summary, detail_rows)
                st.session_state["ai_commentary"] = commentary

            if "ai_commentary" in st.session_state:
                st.markdown(f"""
<div style="background:#ffffff;border:1px solid #c8e6c9;border-radius:10px;padding:18px 22px;margin-top:8px;">
  <div style="font-size:10px;color:#1b5e20;font-weight:700;letter-spacing:0.08em;margin-bottom:10px;">AI COMMENTARY</div>
  <div style="font-size:12px;color:#333;line-height:1.8;white-space:pre-wrap;">{st.session_state["ai_commentary"]}</div>
</div>
""", unsafe_allow_html=True)

        with col2:
            st.markdown("#### 자연어 질문")
            st.caption("포트폴리오 데이터에 대해 자유롭게 질문하세요.")
            question = st.text_input("질문 입력", placeholder="예: MOIC 2배 넘는 회사? / 바이오 섹터 현황?")
            if st.button("질문하기") and question:
                with st.spinner("답변 생성 중..."):
                    answer = answer_question(result_df, question)
                st.session_state["ai_answer"] = answer
                st.session_state["ai_question"] = question

            if "ai_answer" in st.session_state:
                st.markdown(f"""
<div style="background:#ffffff;border:1px solid #e5e5e5;border-radius:10px;padding:18px 22px;margin-top:8px;">
  <div style="font-size:10px;color:#999;font-weight:600;margin-bottom:6px;">Q. {st.session_state.get("ai_question","")}</div>
  <div style="font-size:12px;color:#333;line-height:1.8;white-space:pre-wrap;">{st.session_state["ai_answer"]}</div>
</div>
""", unsafe_allow_html=True)

